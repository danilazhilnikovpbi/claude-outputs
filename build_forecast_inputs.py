"""
build_forecast_inputs.py
========================
Запускается ОДИН РАЗ (или при обновлении данных).

Читает prolongation_test CSV → полная калибровка →
пишет forecast_inputs_2026.xlsx со всеми параметрами модели v3.

Пользователь редактирует Excel → перезапускает forecast_2026_v3.py.

Листы:
  README               — инструкции
  plan_counts          — PRIMARY_PLAN_COUNTS (dims × months), жёлтые ячейки
  package_dist         — вероятности пакетов из вторичных продаж Jul-Dec 2025
  renewal_price        — median price/lesson из вторичных Jul-Dec 2025 + рост цен
  retention_by_renewal — retention rate по payment_no 1-15
  shares               — доли сегментов Present/Earlier/Reanim/Upgrades
  ext_curve            — лаговая кривая (full vs recent + override)
  rates                — flat rate по dim (fallback)
"""

import sys; sys.stdout.reconfigure(encoding='utf-8')
import os
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
DATA_PATH = (
    r"C:\Users\danil\AppData\Local\Temp\prolongation_test_2026-06-08.csv"
)
DATA_CUTOFF_STR = "2026-06-07"
OUT_INPUTS = (
    r"C:\Users\danil\Desktop\Financial Model Project"
    r"\secondary context\forecast_inputs_2026.xlsx"
)

# ── Config (mirrors forecast_2026_v3.py) ──────────────────────────────────────
CAL_WINDOW             = ["2025-07","2025-08","2025-09","2025-10","2025-11","2025-12"]
# Q3 2025 = price base; quarterly_growth = 3%
PRICE_BASE_QUARTER     = 3     # Q3 2025
PRICE_BASE_YEAR        = 2025
PRICE_GROWTH_PCT       = 0.03

FORECAST_MONTHS        = [f"2026-{m:02d}" for m in range(1, 13)]
CURVE_CAL_START        = "2023-01"
CURVE_CAL_START_RECENT = "2024-01"
CURVE_CAL_END          = "2025-06"
CURVE_MAX_LAG          = 36
CURVE_MIN_COHORTS      = 3
CURVE_MIN_COHORTS_RECENT = 2
MIN_N_RATE             = 15
MIN_N_PKG              = 20
MIN_N_PKG_DIST         = 30

DIMS = [
    ('Base','Private'), ('Base','Premium'), ('Base','Group'),
    ('MT',  'Private'), ('MT',  'Premium'), ('MT',  'Group'),
]
SEGS = ['Present', 'Earlier', 'Reanim', 'Upgrades']

PKG_BUCKETS = [8, 10, 16, 20, 24, 32, 40, 56, 64, 80, 120, 128]

def map_to_bucket(pkg):
    """Map a package size to the nearest bucket in PKG_BUCKETS."""
    if pd.isna(pkg) or pkg <= 0:
        return None
    return min(PKG_BUCKETS, key=lambda b: abs(b - pkg))

# Hardcoded fallbacks (matching forecast_2026_v3.py)
PKG_DIST_FALLBACK_BASE_GROUP = {
    8:0.25, 10:0.14, 32:0.13, 16:0.13, 20:0.09, 4:0.08, 36:0.07, 40:0.06
}
PPL_HARDCODED = {
    ('Base','Group', 8): 7.53, ('Base','Group',10): 7.73,
    ('Base','Group',16): 7.06, ('Base','Group',32): 6.62,
    ('Base','Private', 8): 13.92, ('Base','Private',10): 14.66,
    ('Base','Private',16): 12.74,
    ('Base','Premium', 8): 9.69, ('Base','Premium',10): 9.96,
    ('Base','Premium',20): 9.31,
}
RETENTION_BY_PNO_DEFAULT = {
    1:0.184, 2:0.369, 3:0.477, 4:0.503, 5:0.616, 6:0.635,
    7:0.65,  8:0.66,  9:0.67,  10:0.68, 11:0.69, 12:0.70,
    13:0.70, 14:0.71, 15:0.72,
}
RETENTION_PNO_HIGH = 0.72
PROL_RATES_DEFAULT = {
    ('Base','Private'): 0.458, ('Base','Premium'): 0.365, ('Base','Group'): 0.446,
    ('MT',  'Private'): 0.659, ('MT',  'Premium'): 0.435, ('MT',  'Group'): 0.435,
}
SHARES_DEFAULT = {
    'Present': 0.497, 'Earlier': 0.310, 'Reanim': 0.108, 'Upgrades': 0.085,
}

# ── Читаем настройки из существующего Excel (если есть) ──────────────────────
_settings_from_excel = {}
if os.path.exists(OUT_INPUTS):
    try:
        _xl_cfg = __import__('openpyxl').load_workbook(OUT_INPUTS, data_only=True)
        if 'settings' in _xl_cfg.sheetnames:
            for row in _xl_cfg['settings'].iter_rows(min_row=3):
                param = str(row[0].value or '').strip()
                val   = row[1].value
                if param and val is not None and not param.startswith('──'):
                    _settings_from_excel[param] = val
        _xl_cfg.close()
    except Exception:
        pass

# ── Применяем все настройки из Excel ДО загрузки данных ──────────────────────
_fs = str(_settings_from_excel.get('forecast_start', '2026-01'))[:7]
_fe = str(_settings_from_excel.get('forecast_end',   '2026-12'))[:7]
FORECAST_MONTHS = [str(p) for p in pd.period_range(_fs, _fe, freq='M')]

# Обновляем data_cutoff и cal_window если заданы пользователем
_dc_override = _settings_from_excel.get('data_cutoff')
if _dc_override:
    DATA_CUTOFF_STR = str(_dc_override)[:10]

_cws_override = _settings_from_excel.get('cal_window_start')
_cwe_override = _settings_from_excel.get('cal_window_end')
if _cws_override and _cwe_override:
    CAL_WINDOW = [str(p) for p in pd.period_range(str(_cws_override)[:7],
                                                    str(_cwe_override)[:7], freq='M')]

print(f"Settings: data_cutoff={DATA_CUTOFF_STR}  cal={CAL_WINDOW[0]}–{CAL_WINDOW[-1]}"
      f"  forecast={FORECAST_MONTHS[0]}–{FORECAST_MONTHS[-1]}")

# Default PRIMARY_PLAN_COUNTS — для 2026 (hardcoded baseline)
_M2026 = [f"2026-{m:02d}" for m in range(1, 13)]
_PLAN_2026 = {
    ('Base','Private'): dict(zip(_M2026, [134,100,118,145,149,204,221,165,144,133,143,166])),
    ('Base','Premium'): dict(zip(_M2026, [294,220,258,317,327,448,486,362,316,293,314,363])),
    ('Base','Group'):   dict(zip(_M2026, [451,338,395,486,502,686,744,555,485,449,481,557])),
    ('MT',  'Private'): dict(zip(_M2026, [109,116,137,154,159,209,155,155,137,123,126,159])),
    ('MT',  'Premium'): dict(zip(_M2026, [127,135,160,180,186,244,181,180,160,143,147,186])),
    ('MT',  'Group'):   dict(zip(_M2026, [127,135,160,180,186,244,181,180,160,143,147,186])),
}

# PRIMARY_PLAN_COUNTS строится для фактического FORECAST_MONTHS
# Для прошлых/текущих месяцев (есть в CSV) — заполнится реальными данными после загрузки
# Для будущих — используем 2026 defaults или последнее известное значение
PRIMARY_PLAN_COUNTS_DEFAULT = {}
for dim in DIMS:
    last_known = list(_PLAN_2026[dim].values())[-1]
    PRIMARY_PLAN_COUNTS_DEFAULT[dim] = {}
    for mstr in FORECAST_MONTHS:
        PRIMARY_PLAN_COUNTS_DEFAULT[dim][mstr] = _PLAN_2026[dim].get(mstr, last_known)

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING & CALIBRATION
# ══════════════════════════════════════════════════════════════════════════════
DATA_CUTOFF = pd.Timestamp(DATA_CUTOFF_STR)
print("Loading data...")

df = pd.read_csv(DATA_PATH, low_memory=False)
df.columns = [c.replace('prolongation_test[','').replace(']','') for c in df.columns]
for col in ['payment_updated_at','prev_next_payment_planned_dt','next_payment_planned_dt']:
    df[col] = pd.to_datetime(df[col], errors='coerce')
df['usd']         = pd.to_numeric(df['Amount in USD'], errors='coerce')
df['paid_M']      = df['payment_updated_at'].dt.to_period('M')
df['prev_plan_M'] = df['prev_next_payment_planned_dt'].dt.to_period('M')
df['planned_M']   = df['next_payment_planned_dt'].dt.to_period('M')
df['is_sec']      = df['cat_payment'].isin(['same','grad','secondary upgrade'])
df['is_prim']     = df['cat_payment'] == 'primary'
df['course_type'] = (df['full_course_name']
                     .str.contains('Math Tutoring', case=False, na=False)
                     .map({True:'MT', False:'Base'}))
gt_col = 'group_type_c' if 'group_type_c' in df.columns else None
if gt_col:
    df['group_type'] = (df[gt_col].fillna('Unknown')
                        .replace({'Private Eng-Math Tut':'Private',
                                  'Mini':'Premium', 'mini':'Premium'}))
    df['group_type'] = df['group_type'].apply(
        lambda x: x if x in ('Private','Premium','Group') else 'Unknown')
else:
    df['group_type'] = 'Unknown'
# Package size — ВСЕГДА из credits_recieved_package, НЕ из Package (Package=999 — заглушка!)
if 'credits_recieved_package' in df.columns:
    df['pkg'] = pd.to_numeric(df['credits_recieved_package'], errors='coerce').abs()
else:
    df['pkg'] = np.nan
if 'payment_no' in df.columns:
    df['payment_no'] = pd.to_numeric(df['payment_no'], errors='coerce')
else:
    df['payment_no'] = np.nan

def classify_segment(row):
    cp = row['cat_payment']
    if cp == 'secondary upgrade': return 'Upgrades'
    if cp not in ('same','grad'):  return None
    pp, pm = row['prev_plan_M'], row['paid_M']
    if pd.isna(pp) or pd.isna(pm): return None
    if pm == pp: return 'Present'
    if pm <  pp: return 'Earlier'
    return 'Reanim'

df = df[df['payment_updated_at'] <= DATA_CUTOFF].copy()
df['segment'] = df.apply(classify_segment, axis=1)
# pkg_bucket — round pkg to nearest bucket
df['pkg_bucket'] = df['pkg'].apply(map_to_bucket)
print(f"  Loaded {len(df):,} rows")

# ── Авто-заполнение plan_counts из ПОЛНОГО CSV (без обрезки по data_cutoff) ────
# КРИТИЧНО: plan_counts — это ВХОДНЫЕ ДАННЫЕ модели (сколько новых студентов пришло).
# Без этого бэктест 2025 использовал бы 2026-дефолты, которые КАТАСТРОФИЧЕСКИ НЕВЕРНЫ:
#   Jan 2025 реально: Ba/Pre=3, MT/Pri=1, Ba/Gro=1019
#   Jan 2026 план:    Ba/Pre=294, MT/Pri=109, Ba/Gro=451
# → MT dims не существовали в Q1 2025! Использование план=109/мес → фантомные secondary
_df_full_for_plan = pd.read_csv(DATA_PATH, low_memory=False)
_df_full_for_plan.columns = [c.replace('prolongation_test[','').replace(']','')
                               for c in _df_full_for_plan.columns]
_df_full_for_plan['payment_updated_at'] = pd.to_datetime(
    _df_full_for_plan['payment_updated_at'], errors='coerce')
_df_full_for_plan['course_type'] = (
    _df_full_for_plan['full_course_name']
    .str.contains('Math Tutoring', case=False, na=False)
    .map({True:'MT', False:'Base'}))
_gt2 = 'group_type_c' if 'group_type_c' in _df_full_for_plan.columns else None
if _gt2:
    _df_full_for_plan['group_type'] = (
        _df_full_for_plan[_gt2].fillna('Unknown')
        .replace({'Private Eng-Math Tut':'Private','Mini':'Premium','mini':'Premium'}))
    _df_full_for_plan['group_type'] = _df_full_for_plan['group_type'].apply(
        lambda x: x if x in ('Private','Premium','Group') else 'Unknown')
else:
    _df_full_for_plan['group_type'] = 'Unknown'

_df_full_for_plan['is_prim_full'] = _df_full_for_plan['cat_payment'] == 'primary'
_fp_full = _df_full_for_plan.sort_values('payment_updated_at').groupby('student_id').first()
_prim_full = _fp_full[_fp_full['cat_payment'] == 'primary'].copy()
_prim_full['cohort_M'] = _prim_full['payment_updated_at'].dt.to_period('M')
_real_prim_counts = _prim_full.groupby(['cohort_M','course_type','group_type']).size()

# Перезаписываем plan_counts для месяцев с реальными данными
# КРИТИЧНО: сначала обнуляем ВСЕ forecast-месяцы которые есть в CSV,
# потом заполняем реальными значениями. Иначе dims с 0 студентами
# остаются с 2026-дефолтами (e.g. MT/Pre Jan-2025 = 127 вместо 0!)
_forecast_months_set = set(FORECAST_MONTHS)
_months_in_csv = set(str(cm) for (cm,_,_) in _real_prim_counts.index
                      if str(cm) in _forecast_months_set)

# Шаг 1: обнуляем все dim×month комбинации для месяцев которые есть в CSV
_overwrite_count = 0
for mstr in _months_in_csv:
    for (ct_, gt_) in DIMS:
        if (ct_, gt_) in PRIMARY_PLAN_COUNTS_DEFAULT and mstr in PRIMARY_PLAN_COUNTS_DEFAULT[(ct_, gt_)]:
            PRIMARY_PLAN_COUNTS_DEFAULT[(ct_, gt_)][mstr] = 0
            _overwrite_count += 1

# Шаг 2: заполняем реальными значениями (sz > 0)
for (cm, ct_, gt_), sz in _real_prim_counts.items():
    mstr = str(cm)
    if (ct_, gt_) in PRIMARY_PLAN_COUNTS_DEFAULT and mstr in PRIMARY_PLAN_COUNTS_DEFAULT[(ct_, gt_)]:
        PRIMARY_PLAN_COUNTS_DEFAULT[(ct_, gt_)][mstr] = int(sz)

_months_with_real_data = {str(cm) for (cm,_,_) in _real_prim_counts.index
                           if str(cm) in FORECAST_MONTHS}
if _months_with_real_data:
    print(f"  Auto-filled plan_counts from FULL CSV (real primary data): "
          f"{len(_months_with_real_data)} months, {_overwrite_count} dim×month cells")
else:
    print(f"  No forecast months have real primary data in CSV → using plan defaults from Excel")

# ── Ext-curve ─────────────────────────────────────────────────────────────────
first_purch = df.sort_values('payment_updated_at').groupby('student_id').first()
sec_all     = df[df['is_sec']].copy()

def _build_ext_curve(cal_start, min_cohorts):
    _cohorts    = pd.period_range(cal_start, CURVE_CAL_END, freq='M')
    _lag_counts = {k: [] for k in range(1, CURVE_MAX_LAG + 1)}
    for cm in _cohorts:
        stu = set(first_purch[
            (first_purch['cat_payment'] == 'primary') &
            (first_purch['payment_updated_at'].dt.to_period('M') == cm)
        ].index)
        if not stu: continue
        c_sec = sec_all[sec_all['student_id'].isin(stu)]
        for k in range(1, CURVE_MAX_LAG + 1):
            tgt = cm + k
            if tgt > pd.Period(CAL_WINDOW[-1]): continue
            _lag_counts[k].append((c_sec['paid_M'] == tgt).sum() / len(stu))
    return {k: float(np.mean(v)) for k, v in _lag_counts.items() if len(v) >= min_cohorts}

ext_curve_full   = _build_ext_curve(CURVE_CAL_START,        CURVE_MIN_COHORTS)
ext_curve_recent = _build_ext_curve(CURVE_CAL_START_RECENT, CURVE_MIN_COHORTS_RECENT)
ext_curve_final  = {}
for k in range(1, CURVE_MAX_LAG + 1):
    if k in ext_curve_recent:   ext_curve_final[k] = ext_curve_recent[k]
    elif k in ext_curve_full:   ext_curve_final[k] = ext_curve_full[k]

# ── Segment shares ────────────────────────────────────────────────────────────
sec_cal = df[
    df['is_sec'] & df['paid_M'].astype(str).isin(CAL_WINDOW) &
    df['segment'].notna() & df['usd'].notna()
].copy()
total_sec_cal = len(sec_cal)

# ── Prolongation rates ────────────────────────────────────────────────────────
# (moved pool_df definition here so sec_cal can use prev_pkg_bucket)
pool_df     = df[df['next_payment_planned_dt'].notna()].copy()
pool_df['planned_M'] = pool_df['next_payment_planned_dt'].dt.to_period('M')
# pkg_bucket already on df, carried into pool_df
pool_cal_df = pool_df[pool_df['planned_M'].astype(str).isin(CAL_WINDOW)]
glob_pool   = len(pool_cal_df)
glob_rate   = total_sec_cal / glob_pool if glob_pool > 0 else 0.20

# ── seg_lag on sec_cal (used for three-pool rate calibration) ──────────────────
sec_cal_for_3pool = df[
    df['is_sec'] & df['paid_M'].astype(str).isin(CAL_WINDOW) &
    df['segment'].notna() & df['usd'].notna()
].copy()
sec_cal_for_3pool['paid_ym']      = sec_cal_for_3pool['paid_M'].apply(
    lambda p: p.year * 12 + p.month if not pd.isna(p) else None)
sec_cal_for_3pool['prev_plan_ym'] = sec_cal_for_3pool['prev_plan_M'].apply(
    lambda p: p.year * 12 + p.month if not pd.isna(p) else None)
sec_cal_for_3pool['seg_lag'] = (
    sec_cal_for_3pool['paid_ym'] - sec_cal_for_3pool['prev_plan_ym']
)  # 0=Present, <0=Earlier, >0=Reanim

# prev_pkg_bucket on sec_cal: for each secondary sale find the pkg_bucket of the
# previous payment (payment_no - 1) by joining to pool_df via student_id + payment_no
if 'student_id' in df.columns and 'payment_no' in df.columns:
    _pkg_lookup = (
        df[df['payment_no'].notna() & df['pkg_bucket'].notna()]
        [['student_id', 'payment_no', 'pkg_bucket']]
        .rename(columns={'payment_no': '_prev_pno', 'pkg_bucket': 'prev_pkg_bucket'})
    )
    sec_cal = sec_cal.copy()
    # sec_cal row: payment_no = n (the actual secondary payment)
    # prev payment is payment_no - 1, which is what drove the pool prediction
    sec_cal['_prev_pno'] = sec_cal['payment_no'] - 1
    sec_cal = sec_cal.merge(
        _pkg_lookup,
        left_on=['student_id', '_prev_pno'],
        right_on=['student_id', '_prev_pno'],
        how='left'
    )
    sec_cal.drop(columns=['_prev_pno'], inplace=True)
else:
    sec_cal['prev_pkg_bucket'] = None

shares = {}; src_shares = {}
for seg in SEGS:
    if total_sec_cal >= 50:
        shares[seg] = len(sec_cal[sec_cal['segment']==seg]) / total_sec_cal
        src_shares[seg] = f'data n={total_sec_cal}'
    else:
        shares[seg] = SHARES_DEFAULT[seg]
        src_shares[seg] = f'hardcoded default (n={total_sec_cal})'
_s = sum(shares.values())
shares = {k: v/_s for k, v in shares.items()}

rates = {}; src_rates = {}
for ct, gt in DIMS:
    p = len(pool_cal_df[(pool_cal_df['course_type']==ct) & (pool_cal_df['group_type']==gt)])
    s = len(sec_cal[(sec_cal['course_type']==ct) & (sec_cal['group_type']==gt)])
    if p >= MIN_N_RATE:
        rates[(ct,gt)] = s/p
        src_rates[(ct,gt)] = f'data pool={p} sec={s}'
    else:
        rates[(ct,gt)] = PROL_RATES_DEFAULT.get((ct,gt), glob_rate)
        src_rates[(ct,gt)] = f'hardcoded default (pool={p}<{MIN_N_RATE})'

# ── Retention by payment_no ───────────────────────────────────────────────────
# Level 2 fallback: (pno, ct, gt) dim-level retention
# Level 1 expanded: (pno, ct, gt, pkg_bucket) where pool >= 15
retention_by_renewal = {}; src_retention = {}
# retention_expanded: (pno, ct, gt, pkg_bucket) -> rate
retention_expanded = {}; src_retention_expanded = {}

has_pno = 'payment_no' in df.columns and df['payment_no'].notna().sum() > 100
if has_pno:
    pool_pno_cal = pool_cal_df[pool_cal_df['payment_no'].notna()].copy()
    sec_pno_cal  = sec_cal[sec_cal['payment_no'].notna()].copy()

    # ── Level 2: (pno) global ────────────────────────────────────────────────
    for pno in range(1, 16):
        pool_n = len(pool_pno_cal[pool_pno_cal['payment_no'] == float(pno)])
        sec_n  = len(sec_pno_cal[sec_pno_cal['payment_no'] == float(pno + 1)])
        if pool_n >= 10:
            rate_v = min(sec_n / pool_n, 1.0)
            retention_by_renewal[pno] = rate_v
            src_retention[pno] = f'data pool={pool_n} sec={sec_n}'
        else:
            retention_by_renewal[pno] = RETENTION_BY_PNO_DEFAULT.get(pno, RETENTION_PNO_HIGH)
            src_retention[pno] = f'hardcoded default (pool={pool_n}<10)'

    # ── Level 1: (pno, ct, gt, pkg_bucket) ──────────────────────────────────
    pool_pno_pkg_cal = pool_pno_cal[pool_pno_cal['pkg_bucket'].notna()].copy()
    sec_pno_pkg_cal  = sec_pno_cal[sec_pno_cal['prev_pkg_bucket'].notna()].copy()

    for pno in range(1, 16):
        for ct, gt in DIMS:
            for pkg_b in PKG_BUCKETS:
                pool_mask = (
                    (pool_pno_pkg_cal['payment_no'] == float(pno)) &
                    (pool_pno_pkg_cal['course_type'] == ct) &
                    (pool_pno_pkg_cal['group_type']  == gt) &
                    (pool_pno_pkg_cal['pkg_bucket']  == pkg_b)
                )
                pool_n_pkg = pool_mask.sum()
                if pool_n_pkg >= 15:
                    sec_mask = (
                        (sec_pno_pkg_cal['payment_no']      == float(pno + 1)) &
                        (sec_pno_pkg_cal['course_type']      == ct) &
                        (sec_pno_pkg_cal['group_type']       == gt) &
                        (sec_pno_pkg_cal['prev_pkg_bucket']  == pkg_b)
                    )
                    sec_n_pkg = sec_mask.sum()
                    rate_v = min(sec_n_pkg / pool_n_pkg, 1.0)
                    retention_expanded[(pno, ct, gt, pkg_b)] = rate_v
                    src_retention_expanded[(pno, ct, gt, pkg_b)] = (
                        f'data pool={pool_n_pkg} sec={sec_n_pkg}'
                    )
else:
    for pno in range(1, 16):
        retention_by_renewal[pno] = RETENTION_BY_PNO_DEFAULT.get(pno, RETENTION_PNO_HIGH)
        src_retention[pno] = 'hardcoded default (no pno data in CSV)'

n_expanded = len(retention_expanded)
print(f"  Retention expanded (pno,ct,gt,pkg): {n_expanded} combinations with pool>=15")

# ── Three-pool rate calibration ────────────────────────────────────────────────
# Calibrated defaults (from diagnostic — used when pool_n < 10)
RATE_PRESENT_DEFAULT = {1:0.1494, 2:0.2636, 3:0.3097, 4:0.3140, 5:0.4225, 6:0.3989,
                        7:0.4495, 8:0.5156, 9:0.6383, 10:0.6111, 11:0.5758, 12:0.5000}
RATE_EARLIER_DEFAULT = {1:0.0580, 2:0.1166, 3:0.1321, 4:0.1731, 5:0.1713, 6:0.2067,
                        7:0.1957, 8:0.0536, 9:0.1111, 10:0.0312, 11:0.1304}
RATE_REANIM_DEFAULT  = {1:0.0274, 2:0.0458, 3:0.0631, 4:0.0626, 5:0.0661, 6:0.0933,
                        7:0.0870, 8:0.0926, 9:0.0714, 10:0.1379, 11:0.1333}

rate_present_3pool = {}
rate_earlier_3pool = {}
rate_reanim_3pool  = {}
src_rate_3pool     = {}

if has_pno and len(sec_cal_for_3pool) > 0:
    pool_3pool = pool_cal_df[pool_cal_df['payment_no'].notna()].copy()
    sec_3pool  = sec_cal_for_3pool[sec_cal_for_3pool['payment_no'].notna()].copy()

    _cal_periods = [pd.Period(m) for m in CAL_WINDOW]

    for pno in range(1, 14):
        # Aggregate counts across all cal months
        n_present = 0; n_earlier = 0; n_reanim = 0
        pool_T_total = 0; pool_T1_total = 0; pool_Tm1_total = 0

        for T in _cal_periods:
            T1  = T + 1
            Tm1 = T - 1

            pool_T   = pool_3pool[(pool_3pool['planned_M'] == T)   & (pool_3pool['payment_no'] == float(pno))]
            pool_T1  = pool_3pool[(pool_3pool['planned_M'] == T1)  & (pool_3pool['payment_no'] == float(pno))]
            pool_Tm1 = pool_3pool[(pool_3pool['planned_M'] == Tm1) & (pool_3pool['payment_no'] == float(pno))]

            sec_T = sec_3pool[
                (sec_3pool['paid_M'] == T) &
                (sec_3pool['payment_no'] == float(pno + 1))
            ]

            pool_T_total   += len(pool_T)
            pool_T1_total  += len(pool_T1)
            pool_Tm1_total += len(pool_Tm1)

            n_present += (sec_T['seg_lag'] == 0).sum()
            n_earlier += (sec_T['seg_lag'] == -1).sum()
            n_reanim  += (sec_T['seg_lag'] == 1).sum()

        # rate_present = sec_present / pool_T
        if pool_T_total >= 10:
            rate_present_3pool[pno] = n_present / pool_T_total
            src_rate_3pool[(pno, 'present')] = f'data pool_T={pool_T_total} sec={n_present}'
        else:
            rate_present_3pool[pno] = RATE_PRESENT_DEFAULT.get(pno, 0.15)
            src_rate_3pool[(pno, 'present')] = f'hardcoded default (pool_T={pool_T_total}<10)'

        # rate_earlier = sec_earlier1 / pool_T1
        if pool_T1_total >= 10:
            rate_earlier_3pool[pno] = n_earlier / pool_T1_total
            src_rate_3pool[(pno, 'earlier')] = f'data pool_T1={pool_T1_total} sec={n_earlier}'
        else:
            rate_earlier_3pool[pno] = RATE_EARLIER_DEFAULT.get(pno, 0.06)
            src_rate_3pool[(pno, 'earlier')] = f'hardcoded default (pool_T1={pool_T1_total}<10)'

        # rate_reanim = sec_reanim1 / pool_Tm1
        if pool_Tm1_total >= 10:
            rate_reanim_3pool[pno] = n_reanim / pool_Tm1_total
            src_rate_3pool[(pno, 'reanim')] = f'data pool_Tm1={pool_Tm1_total} sec={n_reanim}'
        else:
            rate_reanim_3pool[pno] = RATE_REANIM_DEFAULT.get(pno, 0.03)
            src_rate_3pool[(pno, 'reanim')] = f'hardcoded default (pool_Tm1={pool_Tm1_total}<10)'
else:
    for pno in range(1, 14):
        rate_present_3pool[pno] = RATE_PRESENT_DEFAULT.get(pno, 0.15)
        rate_earlier_3pool[pno] = RATE_EARLIER_DEFAULT.get(pno, 0.06)
        rate_reanim_3pool[pno]  = RATE_REANIM_DEFAULT.get(pno, 0.03)
        src_rate_3pool[(pno, 'present')] = 'hardcoded default (no pno data)'
        src_rate_3pool[(pno, 'earlier')] = 'hardcoded default (no pno data)'
        src_rate_3pool[(pno, 'reanim')]  = 'hardcoded default (no pno data)'

print(f"  Three-pool rates calibrated: pno1 present={rate_present_3pool.get(1,0):.4f} "
      f"earlier={rate_earlier_3pool.get(1,0):.4f} reanim={rate_reanim_3pool.get(1,0):.4f}")

# ── Package distribution ──────────────────────────────────────────────────────
sec_cal_pkg = sec_cal[sec_cal['pkg'].notna() & (sec_cal['pkg'] > 0) & (sec_cal['pkg'] <= 500)].copy()
pkg_dist_final = {}; pkg_dist_src = {}

for ct, gt in DIMS:
    dim_sec = sec_cal_pkg[(sec_cal_pkg['course_type']==ct) & (sec_cal_pkg['group_type']==gt)]
    n_dim   = len(dim_sec)
    if n_dim >= MIN_N_PKG_DIST:
        counts = dim_sec['pkg'].value_counts()
        for pkg_sz, cnt in counts.items():
            pkg_dist_final[(ct, gt, int(pkg_sz))] = cnt / n_dim
            pkg_dist_src[(ct, gt, int(pkg_sz))] = f'data n={n_dim}'
    else:
        fallback_ct = 'Base'
        fallback_gt = gt if gt != 'Unknown' else 'Group'
        base_sec = sec_cal_pkg[
            (sec_cal_pkg['course_type'] == fallback_ct) &
            (sec_cal_pkg['group_type']  == fallback_gt)
        ]
        if len(base_sec) >= MIN_N_PKG_DIST:
            counts = base_sec['pkg'].value_counts()
            total  = len(base_sec)
            for pkg_sz, cnt in counts.items():
                pkg_dist_final[(ct, gt, int(pkg_sz))] = cnt / total
                pkg_dist_src[(ct, gt, int(pkg_sz))] = f'Base/{fallback_gt} fallback n={total}'
        else:
            for pkg_sz, prob in PKG_DIST_FALLBACK_BASE_GROUP.items():
                pkg_dist_final[(ct, gt, pkg_sz)] = prob
                pkg_dist_src[(ct, gt, pkg_sz)] = 'hardcoded Base/Group fallback'

# Normalize per dim
for ct, gt in DIMS:
    dim_keys   = [k for k in pkg_dist_final if k[0]==ct and k[1]==gt]
    total_prob = sum(pkg_dist_final[k] for k in dim_keys)
    if total_prob > 0:
        for k in dim_keys:
            pkg_dist_final[k] /= total_prob

# ── Price per lesson ──────────────────────────────────────────────────────────
# Cal window Q3 only for base period: Jul-Sep 2025
# Ценовая база = Q4-2025 (Oct-Dec). Диагностика: Q3→Q4 рост +15%, Q4→Q1 +3.3%
# Используем Q4 как базу для pgf чтобы Q1-2026 прогноз был точным
_base_cal_window = _settings_from_excel.get('price_base_quarter')
if _base_cal_window == 3:
    cal_q4 = ["2025-07","2025-08","2025-09"]   # Q3 (legacy)
else:
    cal_q4 = ["2025-10","2025-11","2025-12"]   # Q4 (default)
sec_q3 = sec_cal_pkg[sec_cal_pkg['paid_M'].astype(str).isin(cal_q4)].copy()

ppl_final = {}; ppl_src = {}

def _quarter_index(year, month):
    return (year - 2025) * 4 + (month - 1) // 3 + 1

BASE_Q_IDX = _quarter_index(PRICE_BASE_YEAR, 7)  # Q3 2025 index

def _effective_ppl(ppl_base, mstr):
    yr  = int(mstr[:4])
    mo  = int(mstr[5:7])
    dq  = _quarter_index(yr, mo) - BASE_Q_IDX
    return ppl_base * (1 + PRICE_GROWTH_PCT) ** dq

for ct, gt in DIMS:
    pkgs_in_dim = set(k[2] for k in pkg_dist_final if k[0]==ct and k[1]==gt)
    for pkg_sz in pkgs_in_dim:
        pkg_sz_int = int(pkg_sz)
        # Try Q3 data first (cleanest base period)
        sub_q3 = sec_q3[(sec_q3['course_type']==ct) & (sec_q3['group_type']==gt) &
                        (sec_q3['pkg']==pkg_sz_int) & sec_q3['usd'].notna()]
        if len(sub_q3) >= MIN_N_PKG:
            ppl_v = float((sub_q3['usd'] / pkg_sz_int).mean())
            ppl_final[(ct, gt, pkg_sz_int)] = ppl_v
            ppl_src[(ct, gt, pkg_sz_int)] = f'Q3-2025 data n={len(sub_q3)}'
            continue
        # Try full cal_window data
        sub_cal = sec_cal_pkg[(sec_cal_pkg['course_type']==ct) & (sec_cal_pkg['group_type']==gt) &
                              (sec_cal_pkg['pkg']==pkg_sz_int) & sec_cal_pkg['usd'].notna()]
        if len(sub_cal) >= MIN_N_PKG:
            ppl_v = float((sub_cal['usd'] / pkg_sz_int).mean())
            ppl_final[(ct, gt, pkg_sz_int)] = ppl_v
            ppl_src[(ct, gt, pkg_sz_int)] = f'Jul-Dec 2025 data n={len(sub_cal)}'
            continue
        # Hardcoded
        hc = PPL_HARDCODED.get((ct, gt, pkg_sz_int))
        if hc is not None:
            ppl_final[(ct, gt, pkg_sz_int)] = hc
            ppl_src[(ct, gt, pkg_sz_int)] = f'hardcoded calibration (n={len(sub_cal)}<{MIN_N_PKG})'
            continue
        # Fallback: same gt, any ct
        sub_gt = sec_cal_pkg[(sec_cal_pkg['group_type']==gt) &
                             (sec_cal_pkg['pkg']==pkg_sz_int) & sec_cal_pkg['usd'].notna()]
        if len(sub_gt) >= MIN_N_PKG:
            ppl_v = float((sub_gt['usd'] / pkg_sz_int).mean())
            ppl_final[(ct, gt, pkg_sz_int)] = ppl_v
            ppl_src[(ct, gt, pkg_sz_int)] = f'gt fallback n={len(sub_gt)}'
            continue
        # Last resort
        ppl_final[(ct, gt, pkg_sz_int)] = 7.0
        ppl_src[(ct, gt, pkg_sz_int)] = f'last resort 7.0 (n={len(sub_cal)})'

print(f"  Calibration complete")
print(f"  Rates: {', '.join(f'{ct}/{gt}={rates[(ct,gt)]:.1%}' for ct,gt in DIMS)}")
print(f"  Shares: {', '.join(f'{s}={shares[s]:.1%}' for s in SEGS)}")
print(f"  Retention pno1={retention_by_renewal[1]:.1%}, pno3={retention_by_renewal[3]:.1%}, pno7={retention_by_renewal.get(7,0):.1%}")
print(f"  Retention expanded (pno,ct,gt,pkg): {len(retention_expanded)} combinations with pool>=15")
print(f"  Package dims calibrated: {len(set((k[0],k[1]) for k in pkg_dist_final))}/{len(DIMS)}")

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════════════
CP = {
    'navy':   'FF1B2A4A', 'dblue':  'FF2E4B7A', 'teal':   'FF00695C',
    'gold':   'FFFFE082', 'lgrey':  'FFF5F7FA', 'white':  'FFFFFFFF',
    'lblue':  'FFE3EEF9', 'lgreen': 'FFE8F5E9', 'loran':  'FFFFF3E0',
    'lpink':  'FFFCE4EC', 'formula':'FFFFFCE3', 'sub':    'FFECEFF1',
    'edit':   'FFFFFDE7',   # bright yellow = editable
    'lock':   'FFE8EAF6',   # light blue = computed / locked
    'teal2':  'FFE0F2F1',
}
DARK = {'navy','dblue','teal'}

def _fill(k):  return PatternFill('solid', fgColor=CP[k])
def _bord():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def wc(ws, r, c, val, bg='white', bold=False, italic=False,
       align='right', size=10, span=None, fmt=None, wrap=False, color=None):
    if span and span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = _fill(bg)
    fc = color or ('FFFFFFFF' if bg in DARK else 'FF212529')
    cell.font = Font(name='Calibri', size=size, bold=bold, italic=italic, color=fc)
    cell.border = _bord()
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fmt: cell.number_format = fmt

wb = Workbook()

# ─── SHEET: README ────────────────────────────────────────────────────────────
ws0 = wb.active
ws0.title = 'README'
ws0.sheet_view.showGridLines = False
ws0.column_dimensions['A'].width = 80

instructions = [
    ('FORECAST INPUTS 2026  —  Model Parameters (forecast_2026_v3.py)', 'navy', True, 14),
    ('', 'white', False, 10),
    ('COLOR LEGEND:', 'dblue', True, 11),
    ('  Yellow (EDIT)  — editable value. The model reads this directly.', 'edit', False, 10),
    ('  Blue   (LOCK)  — computed from data. Change only if you know why.', 'lock', False, 10),
    ('  Grey           — metadata / source notes.', 'lgrey', False, 10),
    ('', 'white', False, 10),
    ('HOW TO USE:', 'dblue', True, 11),
    ('  1. Run build_forecast_inputs.py  ->  creates this file with current calibration', 'lgrey', False, 10),
    ('  2. Edit yellow cells as needed for your scenario', 'lgrey', False, 10),
    ('  3. Save the Excel file', 'lgrey', False, 10),
    ('  4. Run forecast_2026_v3.py  ->  uses your overrides instead of auto-calibration', 'lgrey', False, 10),
    ('', 'white', False, 10),
    ('SHEETS:', 'dblue', True, 11),
    ('  plan_counts          — new primary students per month Jan-Dec 2026 (wide table)', 'lgrey', False, 10),
    ('  package_dist         — package size probabilities from secondary sales Jul-Dec 2025', 'lgrey', False, 10),
    ('  renewal_price        — price per lesson (base Q3-2025) + quarterly growth factor', 'lgrey', False, 10),
    ('  retention_by_renewal — retention by (pno, ct, gt, pkg_bucket) + (pno, ALL) fallback rows', 'lgrey', False, 10),
    ('  shares               — segment shares Present/Earlier/Reanim/Upgrades', 'lgrey', False, 10),
    ('  ext_curve            — lag curve for future cohort projections (Jun-Dec 2026)', 'lgrey', False, 10),
    ('  rates                — flat prolongation rate by dim (fallback when no pno data)', 'lgrey', False, 10),
    ('', 'white', False, 10),
    ('POOL MODE LOGIC (forecast_2026_v3.py):', 'dblue', True, 11),
    ('  Jan-Apr 2026: mode=data   -> pool_by_pno * retention_by_renewal[pno]', 'lgrey', False, 10),
    ('  May 2026:     mode=blend  -> same * (31/28) to scale 3 missing days', 'lgrey', False, 10),
    ('  Jun-Dec 2026: mode=ext    -> Σ cohort_size * ext_curve[lag]', 'lgrey', False, 10),
    ('', 'white', False, 10),
    ('REVENUE FORMULA:', 'dblue', True, 11),
    ('  Rev = Σ_seg Σ_pkg  total_sec * share[seg] * pkg_dist[dim,pkg]', 'lgrey', False, 10),
    ('              * pkg_size * price_per_lesson[dim,pkg] * price_growth_factor(month)', 'lgrey', False, 10),
    ('  price_growth_factor = (1 + 3%)^(quarters since Q3-2025)', 'lgrey', False, 10),
]
for i, (text, bg, bold, sz) in enumerate(instructions, 1):
    wc(ws0, i, 1, text, bg, bold=bold, align='left', size=sz)
    ws0.row_dimensions[i].height = 18 if sz >= 11 else 16

# ─── SHEET: plan_counts ───────────────────────────────────────────────────────
ws1 = wb.create_sheet('plan_counts')
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 18
for ci in range(2, 15):
    ws1.column_dimensions[get_column_letter(ci)].width = 9

wc(ws1, 1, 1,
   'Primary Student Plan Counts (Jan-Dec 2026)  —  EDIT yellow cells',
   'navy', bold=True, align='left', size=12, span=14)
wc(ws1, 2, 1, 'Dim \\ Month', 'dblue', bold=True, align='center')
for ci, mstr in enumerate(FORECAST_MONTHS, 2):
    wc(ws1, 2, ci, mstr[5:], 'dblue', bold=True, align='center', size=9)
wc(ws1, 2, 14, 'Year Total', 'dblue', bold=True, align='center', size=9)
ws1.row_dimensions[1].height = 26
ws1.row_dimensions[2].height = 18

for ri, (ct, gt) in enumerate(DIMS, 3):
    vals = list(PRIMARY_PLAN_COUNTS_DEFAULT[(ct, gt)].values())
    wc(ws1, ri, 1, f'{ct}/{gt}', 'lgrey', bold=True, align='left')
    for ci, v in enumerate(vals, 2):
        wc(ws1, ri, ci, v, 'edit', fmt='#,##0', align='center')
    wc(ws1, ri, 14, sum(vals), 'lock', fmt='#,##0', bold=True, align='center', italic=True)
    ws1.row_dimensions[ri].height = 18

tr = len(DIMS) + 3
wc(ws1, tr, 1, 'TOTAL (all dims)', 'gold', bold=True, align='left')
for ci, mstr in enumerate(FORECAST_MONTHS, 2):
    total = sum(PRIMARY_PLAN_COUNTS_DEFAULT[(ct,gt)][mstr] for ct,gt in DIMS)
    wc(ws1, tr, ci, total, 'gold', fmt='#,##0', bold=True, align='center')
yr_total = sum(sum(v.values()) for v in PRIMARY_PLAN_COUNTS_DEFAULT.values())
wc(ws1, tr, 14, yr_total, 'gold', fmt='#,##0', bold=True, align='center')
ws1.row_dimensions[tr].height = 20

wc(ws1, tr+2, 1,
   'Edit yellow cells to change the primary student plan, then re-run forecast_2026_v3.py',
   'sub', italic=True, align='left', size=9, span=14)

# ─── SHEET: package_dist ─────────────────────────────────────────────────────
ws_pd = wb.create_sheet('package_dist')
ws_pd.sheet_view.showGridLines = False
for c, w in {1:18, 2:10, 3:14, 4:40}.items():
    ws_pd.column_dimensions[get_column_letter(c)].width = w

wc(ws_pd, 1, 1,
   'Package Distribution  —  secondary sales Jul-Dec 2025  |  EDIT yellow cells to override',
   'navy', bold=True, align='left', size=12, span=4)
for c, h in zip([1,2,3,4], ['Dim', 'Package', 'Probability', 'Note / Source']):
    wc(ws_pd, 2, c, h, 'dblue', bold=True, align='center', size=9)
ws_pd.row_dimensions[1].height = 26
ws_pd.row_dimensions[2].height = 18

_dim_bg_seq = ['lblue','lgreen','loran','teal2','sub','lgrey']
row_pd = 3
for di, (ct, gt) in enumerate(DIMS):
    bg = _dim_bg_seq[di % len(_dim_bg_seq)]
    dim_pkgs = sorted([(k[2], v) for k, v in pkg_dist_final.items() if k[0]==ct and k[1]==gt],
                      key=lambda x: -x[1])
    for pkg_sz, prob in dim_pkgs:
        src_note = pkg_dist_src.get((ct, gt, pkg_sz), '')
        wc(ws_pd, row_pd, 1, f'{ct}/{gt}',  bg,     align='left',   size=9)
        wc(ws_pd, row_pd, 2, pkg_sz,         bg,     fmt='#,##0',   align='center', size=9)
        wc(ws_pd, row_pd, 3, round(prob, 4), 'edit', fmt='0.0000',  align='center')
        wc(ws_pd, row_pd, 4, src_note,       'lgrey', align='left', size=8, italic=True)
        ws_pd.row_dimensions[row_pd].height = 15
        row_pd += 1

wc(ws_pd, row_pd+1, 1,
   'Probabilities are auto-normalized per dim in the model. Edit to override distribution.',
   'sub', italic=True, align='left', size=9, span=4)
wc(ws_pd, row_pd+2, 1,
   f'MT dims with thin data use Base/{{}}-fallback. Base/Group hardcoded fallback: {PKG_DIST_FALLBACK_BASE_GROUP}',
   'sub', italic=True, align='left', size=8, span=4, wrap=True)

# ─── SHEET: renewal_price ─────────────────────────────────────────────────────
ws_rp = wb.create_sheet('renewal_price')
ws_rp.sheet_view.showGridLines = False
for c, w in {1:18, 2:10, 3:16, 4:14, 5:16, 6:40}.items():
    ws_rp.column_dimensions[get_column_letter(c)].width = w

wc(ws_rp, 1, 1,
   'Renewal Price per Lesson (base Q3-2025)  —  EDIT yellow cells to override',
   'navy', bold=True, align='left', size=12, span=6)
for c, h in zip([1,2,3,4,5,6],
                ['Dim', 'Package', 'Price/Lesson (base)', 'Growth%/quarter',
                 'Effective Jan-2026', 'Note / Source']):
    wc(ws_rp, 2, c, h, 'dblue', bold=True, align='center', size=9)
ws_rp.row_dimensions[1].height = 26
ws_rp.row_dimensions[2].height = 18

row_rp = 3
q_jan26 = _quarter_index(2026, 1) - BASE_Q_IDX
pgf_jan = (1 + PRICE_GROWTH_PCT) ** q_jan26

for di, (ct, gt) in enumerate(DIMS):
    bg = _dim_bg_seq[di % len(_dim_bg_seq)]
    dim_pkgs = sorted([(k[2], v) for k, v in ppl_final.items() if k[0]==ct and k[1]==gt],
                      key=lambda x: x[0])
    for pkg_sz, ppl_v in dim_pkgs:
        src_note = ppl_src.get((ct, gt, pkg_sz), '')
        eff_jan  = round(ppl_v * pgf_jan, 4)
        wc(ws_rp, row_rp, 1, f'{ct}/{gt}',         bg,      align='left',      size=9)
        wc(ws_rp, row_rp, 2, pkg_sz,                bg,      fmt='#,##0',      align='center', size=9)
        wc(ws_rp, row_rp, 3, round(ppl_v, 4),       'edit',  fmt='$#,##0.0000', align='center')
        wc(ws_rp, row_rp, 4, PRICE_GROWTH_PCT,      'lock',  fmt='0.0%',       align='center', italic=True)
        wc(ws_rp, row_rp, 5, eff_jan,               'lock',  fmt='$#,##0.0000', align='center', italic=True)
        wc(ws_rp, row_rp, 6, src_note,              'lgrey', align='left',     size=8, italic=True)
        ws_rp.row_dimensions[row_rp].height = 15
        row_rp += 1

wc(ws_rp, row_rp+1, 1,
   f'Base period: Q3-2025 (Jul-Sep). Growth factor = (1+3%)^(quarters from Q3-2025). '
   f'Jan-2026 is Q1-2026 = 2 quarters after Q3-2025, so pgf = {pgf_jan:.4f}.',
   'sub', italic=True, align='left', size=9, span=6, wrap=True)
ws_rp.row_dimensions[row_rp+1].height = 30

# ─── SHEET: retention_by_renewal — long-format with pkg_bucket breakdown ──────
ws_ret = wb.create_sheet('retention_by_renewal')
ws_ret.sheet_view.showGridLines = False
for c, w in {1:8, 2:14, 3:14, 4:12, 5:12, 6:12, 7:14, 8:40}.items():
    ws_ret.column_dimensions[get_column_letter(c)].width = w

wc(ws_ret, 1, 1,
   'Retention by (pno, ct, gt, pkg_bucket)  —  Override column takes priority. '
   'Fallback: pkg_bucket=ALL (dim-level), then global pno default.',
   'navy', bold=True, align='left', size=11, span=8)
wc(ws_ret, 2, 1,
   'Fallback hierarchy: Level1=(pno,ct,gt,pkg) pool>=15  →  Level2=(pno,ct,gt) ALL  →  Level3=global(pno)',
   'sub', italic=True, align='left', size=9, span=8)
for c, h in zip(range(1, 9),
                ['pno', 'ct', 'gt', 'pkg_bucket', 'Retention', 'Override', 'n_pool', 'Note / Source']):
    wc(ws_ret, 3, c, h, 'dblue', bold=True, align='center', size=9)
ws_ret.row_dimensions[1].height = 26
ws_ret.row_dimensions[2].height = 16
ws_ret.row_dimensions[3].height = 18

_pno_bg = ['lblue','lgreen','loran','teal2','sub','lgrey']

def _parse_pool_n(src_v):
    """Extract pool count from source string like 'data pool=42 sec=18'."""
    if 'pool=' not in src_v:
        return '--'
    try:
        for part in src_v.split():
            if part.startswith('pool='):
                return int(part.split('=')[1])
    except Exception:
        pass
    return '--'

ret_row = 4

# ── Part 1: expanded rows (pno, ct, gt, pkg_bucket) with Level-1 data ────────
for pno in range(1, 16):
    for ct, gt in DIMS:
        for pkg_b in PKG_BUCKETS:
            key = (pno, ct, gt, pkg_b)
            if key not in retention_expanded:
                continue
            rate_v = retention_expanded[key]
            src_v  = src_retention_expanded.get(key, '')
            bg = _pno_bg[(pno - 1) % len(_pno_bg)]
            wc(ws_ret, ret_row, 1, pno,              bg,     fmt='#,##0', align='center', size=9)
            wc(ws_ret, ret_row, 2, ct,               bg,     align='left', size=9)
            wc(ws_ret, ret_row, 3, gt,               bg,     align='left', size=9)
            wc(ws_ret, ret_row, 4, pkg_b,            bg,     fmt='#,##0', align='center', size=9)
            wc(ws_ret, ret_row, 5, round(rate_v, 4), 'lock', fmt='0.0000', align='center')
            wc(ws_ret, ret_row, 6, None,             'edit', fmt='0.0000', align='center')
            wc(ws_ret, ret_row, 7, _parse_pool_n(src_v),
               bg, fmt='#,##0' if isinstance(_parse_pool_n(src_v), int) else None,
               align='center', size=9)
            wc(ws_ret, ret_row, 8, src_v,            'lgrey', align='left', size=8, italic=True)
            ws_ret.row_dimensions[ret_row].height = 14
            ret_row += 1

# ── Part 2: dim-level fallback rows (pno, ct, gt, 'ALL') ─────────────────────
# These are Level-2 fallback: (pno, ct, gt) with no pkg split
for pno in range(1, 16):
    rate_v = retention_by_renewal.get(pno, RETENTION_BY_PNO_DEFAULT.get(pno, RETENTION_PNO_HIGH))
    src_v  = src_retention.get(pno, '')
    bg = _pno_bg[(pno - 1) % len(_pno_bg)]
    # Write one 'ALL' row per pno (global, covers all dims as fallback)
    wc(ws_ret, ret_row, 1, pno,              'gold', fmt='#,##0', bold=True, align='center', size=9)
    wc(ws_ret, ret_row, 2, 'ALL',            'gold', align='center', size=9, bold=True)
    wc(ws_ret, ret_row, 3, 'ALL',            'gold', align='center', size=9, bold=True)
    wc(ws_ret, ret_row, 4, 'ALL',            'gold', align='center', size=9, bold=True)
    wc(ws_ret, ret_row, 5, round(rate_v, 4), 'edit', fmt='0.0000', align='center')
    wc(ws_ret, ret_row, 6, None,             'edit', fmt='0.0000', align='center')
    wc(ws_ret, ret_row, 7, _parse_pool_n(src_v),
       'gold', fmt='#,##0' if isinstance(_parse_pool_n(src_v), int) else None,
       align='center', size=9)
    wc(ws_ret, ret_row, 8, src_v + '  ← Level-2 fallback (ALL pkg)',
       'lgrey', align='left', size=8, italic=True)
    ws_ret.row_dimensions[ret_row].height = 15
    ret_row += 1

last_ret = ret_row
wc(ws_ret, last_ret+1, 1,
   f'pno=1 = primary (not secondary). pno=2 = first renewal. '
   f'Override (col F) takes priority over Retention (col E) if non-empty. '
   f'pkg_bucket rows = Level-1 (pool>=15). ALL rows = Level-2 fallback. '
   f'For pno>=16 model uses RETENTION_PNO_HIGH={RETENTION_PNO_HIGH:.1%}.',
   'sub', italic=True, align='left', size=9, span=8, wrap=True)
ws_ret.row_dimensions[last_ret+1].height = 36

# ── THREE-POOL RATES section ─────────────────────────────────────────────────
# Historical segment shares used as calibration baseline
HIST_SHARES_3POOL = {'Present': 0.497, 'Earlier': 0.310, 'Reanim': 0.108}

sep_row = last_ret + 3
wc(ws_ret, sep_row, 1,
   '━━━━━━━━  THREE-POOL RATES  ━━━━━━━━  '
   f'Hist shares: Present={HIST_SHARES_3POOL["Present"]:.3f}  '
   f'Earlier={HIST_SHARES_3POOL["Earlier"]:.3f}  '
   f'Reanim={HIST_SHARES_3POOL["Reanim"]:.3f}  '
   '—  Edit rates below; scaling applied by forecast_2026_v3.py based on shares sheet',
   'dblue', bold=True, align='left', size=10, span=8)
ws_ret.row_dimensions[sep_row].height = 22

hdr_row = sep_row + 1
for c, h in zip(range(1, 9),
                ['pno', 'rate_present', 'rate_earlier_1', 'rate_reanim_1',
                 'note_present', 'note_earlier', 'note_reanim', 'hist_shares']):
    wc(ws_ret, hdr_row, c, h, 'dblue', bold=True, align='center', size=9)
ws_ret.row_dimensions[hdr_row].height = 18

data_row = hdr_row + 1
for pno in range(1, 14):
    wc(ws_ret, data_row, 1, pno,
       'gold', fmt='#,##0', bold=True, align='center', size=9)
    wc(ws_ret, data_row, 2, round(rate_present_3pool.get(pno, 0.15), 4),
       'edit', fmt='0.0000', align='center')
    wc(ws_ret, data_row, 3, round(rate_earlier_3pool.get(pno, 0.06), 4),
       'edit', fmt='0.0000', align='center')
    wc(ws_ret, data_row, 4, round(rate_reanim_3pool.get(pno, 0.03), 4),
       'edit', fmt='0.0000', align='center')
    wc(ws_ret, data_row, 5, src_rate_3pool.get((pno, 'present'), ''),
       'lgrey', align='left', size=8, italic=True)
    wc(ws_ret, data_row, 6, src_rate_3pool.get((pno, 'earlier'), ''),
       'lgrey', align='left', size=8, italic=True)
    wc(ws_ret, data_row, 7, src_rate_3pool.get((pno, 'reanim'), ''),
       'lgrey', align='left', size=8, italic=True)
    wc(ws_ret, data_row, 8,
       f'Present_hist={HIST_SHARES_3POOL["Present"]}  '
       f'Earlier_hist={HIST_SHARES_3POOL["Earlier"]}  '
       f'Reanim_hist={HIST_SHARES_3POOL["Reanim"]}',
       'lgrey', align='left', size=8, italic=True)
    ws_ret.row_dimensions[data_row].height = 15
    data_row += 1

wc(ws_ret, data_row + 1, 1,
   'THREE-POOL section is read by forecast_2026_v3.py for data/blend mode months. '
   'rate_present = pool[T]×rate→sec; rate_earlier = pool[T+1]×rate→early-pay; '
   'rate_reanim = pool[T-1]×rate→late-pay. Effective rates are scaled by '
   '(current_share / hist_share) from the shares sheet.',
   'sub', italic=True, align='left', size=8, span=8, wrap=True)
ws_ret.row_dimensions[data_row + 1].height = 40

# ─── SHEET: shares ────────────────────────────────────────────────────────────
ws_sh = wb.create_sheet('shares')
ws_sh.sheet_view.showGridLines = False
for c, w in {1:16, 2:16, 3:40}.items():
    ws_sh.column_dimensions[get_column_letter(c)].width = w

wc(ws_sh, 1, 1,
   f'Segment Shares  —  EDIT yellow column (must sum to 1.0 across all 4 segments)',
   'navy', bold=True, align='left', size=12, span=3)
for c, h in zip([1,2,3], ['Segment', 'Share', 'Note / Source']):
    wc(ws_sh, 2, c, h, 'dblue', bold=True, align='center', size=9)
ws_sh.row_dimensions[1].height = 26
ws_sh.row_dimensions[2].height = 18

seg_bgs = ['lblue','lgreen','loran','sub']
seg_desc = {
    'Present':  'Paid in expected month (paid_M == prev_plan_M)',
    'Earlier':  'Paid before expected month (paid_M < prev_plan_M)',
    'Reanim':   'Paid after expected month (paid_M > prev_plan_M)',
    'Upgrades': 'cat_payment = secondary upgrade',
}
for ri, (seg, bg) in enumerate(zip(SEGS, seg_bgs), 3):
    wc(ws_sh, ri, 1, seg,                     bg,     bold=True, align='left')
    wc(ws_sh, ri, 2, round(shares[seg], 4),   'edit', fmt='0.0000', align='center')
    wc(ws_sh, ri, 3, f'{src_shares[seg]} | {seg_desc[seg]}',
       'lgrey', align='left', size=8, italic=True)
    ws_sh.row_dimensions[ri].height = 18

tr_sh = len(SEGS) + 3
wc(ws_sh, tr_sh, 1, 'SUM', 'gold', bold=True, align='left')
wc(ws_sh, tr_sh, 2, round(sum(shares.values()), 4), 'gold', fmt='0.0000', bold=True, align='center')
wc(ws_sh, tr_sh, 3, f'Calibrated from n={total_sec_cal} secondary sales {CAL_WINDOW[0]}-{CAL_WINDOW[-1]}',
   'lgrey', align='left', size=9)

wc(ws_sh, tr_sh+2, 1,
   'Edit yellow cells. The model normalizes shares automatically if sum != 1.0.',
   'sub', italic=True, align='left', size=9, span=3)

# ─── SHEET: ext_curve ─────────────────────────────────────────────────────────
ws_ec = wb.create_sheet('ext_curve')
ws_ec.sheet_view.showGridLines = False
for c, w in {1:10, 2:14, 3:14, 4:14, 5:12, 6:12}.items():
    ws_ec.column_dimensions[get_column_letter(c)].width = w

wc(ws_ec, 1, 1,
   'Ext-curve (Lag Curve for Future Cohorts)  —  EDIT col E (Override) only if needed',
   'navy', bold=True, align='left', size=12, span=6)
recent_max = max(ext_curve_recent.keys()) if ext_curve_recent else '?'
full_max   = max(ext_curve_final.keys())  if ext_curve_final  else '?'
wc(ws_ec, 2, 1,
   f'Recent (2024-01 to {CURVE_CAL_END}): lags 1-{recent_max}  |  '
   f'Full fallback (2023-01 to {CURVE_CAL_END}): lags {int(recent_max)+1 if recent_max != "?" else "?"}-{full_max}  |  '
   f'Used by mode=ext (Jun-Dec 2026)',
   'sub', italic=True, align='left', size=9, span=6)
for c, h in zip([1,2,3,4,5,6],
                ['Lag', 'Full (2023+)', 'Recent (2024+)', 'Used (auto)', 'Override', 'Source']):
    wc(ws_ec, 3, c, h, 'dblue', bold=True, align='center', size=9)
ws_ec.row_dimensions[1].height = 26
ws_ec.row_dimensions[2].height = 16
ws_ec.row_dimensions[3].height = 18

for ri, k in enumerate(sorted(ext_curve_final.keys()), 4):
    full_v   = ext_curve_full.get(k)
    recent_v = ext_curve_recent.get(k)
    used_v   = ext_curve_final[k]
    src_tag  = 'recent' if k in ext_curve_recent else 'full'
    bg_row   = 'lgreen' if src_tag == 'recent' else 'loran'
    wc(ws_ec, ri, 1, f'lag {k}',                     'lgrey',   align='center', size=9)
    wc(ws_ec, ri, 2, full_v   if full_v   else '--',  'lock' if full_v   else 'lgrey',
       fmt='0.000%' if full_v else None,   align='center', size=9)
    wc(ws_ec, ri, 3, recent_v if recent_v else '--',  'lock' if recent_v else 'lgrey',
       fmt='0.000%' if recent_v else None, align='center', size=9)
    wc(ws_ec, ri, 4, used_v,                           bg_row,    fmt='0.000%',   align='center', bold=True, size=9)
    wc(ws_ec, ri, 5, '',                               'edit',    fmt='0.000%',   align='center', size=9)
    wc(ws_ec, ri, 6, src_tag,                          'lgrey',   align='center', size=8, italic=True)
    ws_ec.row_dimensions[ri].height = 14

last_ec = len(ext_curve_final) + 4
wc(ws_ec, last_ec+1, 1,
   'Col E (Override): if filled, model uses it instead of auto value. '
   'Recent curve is preferred for lags 1-16; full curve covers longer lags from 2023 cohorts.',
   'sub', italic=True, align='left', size=9, span=6, wrap=True)
ws_ec.row_dimensions[last_ec+1].height = 30

# ─── SHEET: rates (flat rate fallback) ────────────────────────────────────────
ws_r = wb.create_sheet('rates')
ws_r.sheet_view.showGridLines = False
for c, w in {1:18, 2:14, 3:14, 4:40}.items():
    ws_r.column_dimensions[get_column_letter(c)].width = w

wc(ws_r, 1, 1,
   'Flat Prolongation Rate by Dim  —  EDIT yellow column  (fallback when no payment_no data)',
   'navy', bold=True, align='left', size=12, span=4)
for c, h in zip([1,2,3,4],
                ['Dim', 'Rate (edit/override)', 'Auto-calibrated', 'Note / Source']):
    wc(ws_r, 2, c, h, 'dblue', bold=True, align='center', size=9)
ws_r.row_dimensions[1].height = 26
ws_r.row_dimensions[2].height = 18

for ri, (ct, gt) in enumerate(DIMS, 3):
    bg = _dim_bg_seq[(ri-3) % len(_dim_bg_seq)]
    r  = rates[(ct, gt)]
    wc(ws_r, ri, 1, f'{ct}/{gt}',       bg,     bold=True, align='left')
    wc(ws_r, ri, 2, round(r, 4),        'edit', fmt='0.0000', align='center')
    wc(ws_r, ri, 3, round(r, 4),        'lock', fmt='0.0000', align='center', italic=True)
    wc(ws_r, ri, 4, src_rates[(ct, gt)],'lgrey', align='left', size=9, italic=True)
    ws_r.row_dimensions[ri].height = 18

tr_r = len(DIMS) + 3
wc(ws_r, tr_r, 1,
   f'Global fallback rate (when pool < {MIN_N_RATE}): {glob_rate:.4f}  '
   f'(pool={glob_pool}, sec={total_sec_cal})',
   'sub', italic=True, align='left', size=9, span=4)
wc(ws_r, tr_r+2, 1,
   'This flat rate is used as a fallback in mode=ext (Jun-Dec 2026) for pool_display calculation. '
   'For mode=data/blend the model uses retention_by_renewal[pno] instead.',
   'sub', italic=True, align='left', size=9, span=4, wrap=True)
ws_r.row_dimensions[tr_r+2].height = 30

# ── SHEET: settings ───────────────────────────────────────────────────────────
ws_cfg = wb.create_sheet('settings')
ws_cfg.sheet_view.showGridLines = False
for c, w in {1:28, 2:22, 3:50}.items():
    ws_cfg.column_dimensions[get_column_letter(c)].width = w

wc(ws_cfg, 1, 1, 'НАСТРОЙКИ МОДЕЛИ  —  меняй жёлтые ячейки, сохрани, перезапусти forecast_2026_v3.py',
   'navy', bold=True, align='left', size=12, span=3)
ws_cfg.row_dimensions[1].height = 28

for c, h in zip([1,2,3], ['Параметр', 'Значение', 'Описание / Как влияет']):
    wc(ws_cfg, 2, c, h, 'dblue', bold=True, align='center', size=9)
ws_cfg.row_dimensions[2].height = 18

# Используем значения из существующего Excel (если они были изменены пользователем)
# Fallback — дефолты. Так пользовательские настройки не сбрасываются при пересборке.
def _cfg(key, default):
    v = _settings_from_excel.get(key)
    return str(v)[:10] if v is not None else default

_settings = [
    # ── Окно данных и прогноза ────────────────────────────────────────────────
    ('── ОКНО ДАННЫХ И ПРОГНОЗА ──', None, '', 'dblue'),
    ('data_cutoff',    _cfg('data_cutoff', DATA_CUTOFF_STR),
     'Дата последних данных. Месяцы ДО этой даты используют режим data/blend. '
     'Например: 2024-12-31 = данные по декабрь 2024, прогноз с января 2025.', 'edit'),
    ('cal_window_start', _cfg('cal_window_start', CAL_WINDOW[0]),
     'Начало окна калибровки рейтов/долей/AOV. Формат: YYYY-MM.', 'edit'),
    ('cal_window_end',   _cfg('cal_window_end', CAL_WINDOW[-1]),
     'Конец окна калибровки. По умолчанию — последний полный месяц до data_cutoff.', 'edit'),
    ('forecast_start', _cfg('forecast_start', str(FORECAST_MONTHS[0])),
     'Первый месяц прогноза. Формат: YYYY-MM.', 'edit'),
    ('forecast_end',   _cfg('forecast_end', str(FORECAST_MONTHS[-1])),
     'Последний месяц прогноза. Например: 2027-05 = прогноз на 12 мес. вперёд от data_cutoff.', 'edit'),
    # ── AOV коррекция ─────────────────────────────────────────────────────────
    ('── КОРРЕКЦИЯ AOV ──', None, '', 'dblue'),
    ('aov_adjustment_factor', float(_settings_from_excel.get('aov_adjustment_factor', 1.10)),
     '❗ ПОЧЕМУ AOV может отличаться от факта:\n'
     '1. Цены calibrated на среднем (mean) — чувствительно к выбросам и сезонности.\n'
     '2. Пакеты pkg=999 (кастомные) исключены из калибровки — реальный AOV выше.\n'
     '3. Распределение пакетов июл-дек 2025 может отличаться от 2026.\n'
     '→ Коэффициент 1.10 = скорректировать выручку на 10% (бэктест: Jan 2026 = -0.2%).\n'
     '→ Поставь 1.0 чтобы без коррекции. Меняй под обновлённые данные.', 'edit'),
    # ── Рост цен ──────────────────────────────────────────────────────────────
    ('── РОСТ ЦЕН ──', None, '', 'dblue'),
    ('price_growth_pct_per_quarter', float(_settings_from_excel.get('price_growth_pct_per_quarter', 0.03)),
     'Рост цены за урок в квартал (3% = +3%/квартал). База = Q3 2025.', 'edit'),
    ('price_base_year',    2025,
     'Год базового периода для роста цен.', 'lock'),
    ('price_base_quarter', int(_settings_from_excel.get('price_base_quarter', 4)),
     'Квартал базового периода (4 = Q4 = окт-дек, рекомендуется). '
     'Диагностика: Q3→Q4 рост +15%, поэтому Q4 = более точная база для 2026.', 'lock'),
    # ── Корректировка retention ────────────────────────────────────────────────
    ('── КОРРЕКТИРОВКА RETENTION ──', None, '', 'dblue'),
    ('retention_adjustment', float(_settings_from_excel.get('retention_adjustment', 1.0)),
     '❗ ТРЕНД СНИЖЕНИЯ RETENTION:\n'
     'May 2026 actual: retention=38.4% vs калибровка Jul-Dec 2025: 43.6% (-12%).\n'
     'Retention постепенно снижается год к году.\n'
     '→ 0.88 = умножить все retention[pno] на 0.88 (-12%)\n'
     '→ 1.0  = без коррекции (использовать 2025 ставки)\n'
     'Рекомендация для 2026 прогноза: попробуй 0.88-0.92.', 'edit'),
]

ri = 3
for item in _settings:
    if len(item) == 4:
        param, val, desc, bg = item
    if bg == 'dblue':
        wc(ws_cfg, ri, 1, param, 'dblue', bold=True, align='left', size=10, span=3)
        ws_cfg.row_dimensions[ri].height = 18
        ri += 1
        continue
    wc(ws_cfg, ri, 1, param,                'lgrey', bold=True, align='left', size=10)
    if val is not None:
        fmt = None
        if isinstance(val, float) and val < 1.5:
            fmt = '0.00'
        elif isinstance(val, float):
            fmt = '0.000'
        wc(ws_cfg, ri, 2, val, 'edit', fmt=fmt, align='center', size=11, bold=True)
    else:
        wc(ws_cfg, ri, 2, '—', 'lgrey', align='center', size=10)
    wc(ws_cfg, ri, 3, desc, 'sub', align='left', size=8, italic=True, wrap=True)
    ws_cfg.row_dimensions[ri].height = max(14, min(60, len(desc) // 3))
    ri += 1

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT_INPUTS)
print(f"\n  Saved: {OUT_INPUTS}")
print("  Sheets: README | plan_counts | package_dist | renewal_price | "
      "retention_by_renewal | shares | ext_curve | rates | settings")
print("\nNext steps:")
print("  1. Open the file, review yellow cells and adjust for your scenario")
print("  2. Run forecast_2026_v3.py  ->  it will load your overrides automatically")
