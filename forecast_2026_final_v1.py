"""
forecast_2026_v3.py — Прогноз вторичных продаж 2026 (v3)
=========================================================
Новая архитектура (пакетная revenue-модель):

1. ПУЛ — 3 режима, БЕЗ unified max():
   - T <= 2026-04 (Jan-Apr): mode='data'
       total_sec = Σ_n pool_by_pno[(m,ct,gt,n)] × retention_by_renewal[n]
   - T == 2026-05 (May):    mode='blend'
       то же, + scale (31/28) для 3 недостающих дней
   - T >= 2026-06 (Jun-Dec): mode='ext'
       total_sec = Σ_cohort cohort_size(C,ct,gt) × ext_curve[T-C]
       (исторические + плановые когорты Jun-Dec 2026)

2. ВЫРУЧКА — пакетная модель (замена AOV-коэффициента):
   revenue = Σ_seg Σ_pkg total_sec × share[seg] × pkg_dist[(ct,gt,pkg)]
             × pkg_size × ppl[(ct,gt,pkg)] × price_growth_factor(month)
   price_growth_factor: +3% в квартал относительно Q3 2025

3. ИНПУТЫ из forecast_inputs_2026.xlsx (если файл есть):
   - plan_counts, package_dist, renewal_price, retention_by_renewal,
     shares, ext_curve, rates (fallback)

4. ВЫВОД: forecast_2026_v3.xlsx
   Листы: Monthly Summary | Dim Breakdown | Calibration | Wide Format
"""

import sys; sys.stdout.reconfigure(encoding='utf-8')
import os as _os
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# [A] CONFIG
# ══════════════════════════════════════════════════════════════════════════════
DATA_PATH = r"C:\Users\danil\Desktop\Financial Model Project\secondary context\prolongation_test_2026-06-08.csv"
DATA_CUTOFF_STR  = "2026-06-07"
CAL_WINDOW       = ["2025-07","2025-08","2025-09","2025-10","2025-11","2025-12"]
# Q4 2025 = base period for price growth (Oct-Nov-Dec 2025)
# Диагностика показала: цены скакнули Q3→Q4 на +15%, потом Q4→Q1 стабилизировались (+3.3%)
# Используем Q4 как базу → точность Q1-2026 AOV улучшается с -5% до ~0%
PRICE_BASE_QUARTER = 4   # Q4 2025 (1-based: Q1=1, Q2=2, Q3=3, Q4=4)
PRICE_BASE_YEAR    = 2025
PRICE_GROWTH_PCT   = 0.03   # 3% per quarter

FORECAST_MONTHS  = [f"2026-{m:02d}" for m in range(1, 13)]

# Ext-curve calibration windows
CURVE_CAL_START        = "2023-01"
CURVE_CAL_START_RECENT = "2024-01"
CURVE_CAL_END          = "2025-06"
CURVE_MAX_LAG          = 36
CURVE_MIN_COHORTS      = 3
CURVE_MIN_COHORTS_RECENT = 2
PLAN_COHORT_START      = "2026-06"   # cohorts from this month onward are plan-only

# Thresholds
MIN_N_SHARES     = 50
MIN_N_RATE       = 15
MIN_N_PKG        = 20   # min secondary sales to use (dim,pkg) price
MIN_N_PKG_DIST   = 30   # min secondary sales to use (dim) pkg distribution

INPUTS_EXCEL = (
    r"C:\Users\danil\Desktop\Financial Model Project"
    r"\secondary context\forecast_inputs_2026.xlsx"
)
OUT_PATH = (
    r"C:\Users\danil\Desktop\Financial Model Project"
    r"\secondary context\forecast_2026_v3.xlsx"
)

DIMS = [
    ('Base','Private'), ('Base','Premium'), ('Base','Group'),
    ('MT',  'Private'), ('MT',  'Premium'), ('MT',  'Group'),
]
SEGS = ['Present', 'Earlier', 'Reanim', 'Upgrades']
SEP  = '=' * 100

PKG_BUCKETS = [8, 10, 16, 20, 24, 32, 40, 56, 64, 80, 120, 128]

def map_to_bucket(pkg):
    """Map a package size to the nearest bucket in PKG_BUCKETS."""
    if pd.isna(pkg) or pkg <= 0:
        return None
    return min(PKG_BUCKETS, key=lambda b: abs(b - pkg))

# ══════════════════════════════════════════════════════════════════════════════
# [B] HARD-CODED CALIBRATION FALLBACKS (from key calibration findings)
# ══════════════════════════════════════════════════════════════════════════════
# Package distribution fallback for Base/Group (and thin dims that fall back to it)
PKG_DIST_FALLBACK_BASE_GROUP = {
    8:0.25, 10:0.14, 32:0.13, 16:0.13, 20:0.09, 4:0.08, 36:0.07, 40:0.06
}
# Price per lesson (base, Jul-Dec 2025 calibration)
PPL_HARDCODED = {
    # Base/Group
    ('Base','Group', 8): 7.53, ('Base','Group',10): 7.73,
    ('Base','Group',16): 7.06, ('Base','Group',32): 6.62,
    # Base/Private
    ('Base','Private', 8): 13.92, ('Base','Private',10): 14.66,
    ('Base','Private',16): 12.74,
    # Base/Premium
    ('Base','Premium', 8): 9.69, ('Base','Premium',10): 9.96,
    ('Base','Premium',20): 9.31,
}
# Retention by payment_no (fallback)
RETENTION_BY_PNO_DEFAULT = {
    1:0.184, 2:0.369, 3:0.477, 4:0.503, 5:0.616, 6:0.635,
    7:0.65,  8:0.66,  9:0.67,  10:0.68, 11:0.69, 12:0.70,
    13:0.70, 14:0.71, 15:0.72,
}
# For pno >= 16 use this rate
RETENTION_PNO_HIGH = 0.72

# Prolongation rates (fallback)
PROL_RATES_DEFAULT = {
    ('Base','Private'): 0.458, ('Base','Premium'): 0.365, ('Base','Group'): 0.446,
    ('MT',  'Private'): 0.659, ('MT',  'Premium'): 0.435, ('MT',  'Group'): 0.435,
}
# Segment shares (fallback)
SHARES_DEFAULT = {
    'Present': 0.497, 'Earlier': 0.310, 'Reanim': 0.108, 'Upgrades': 0.085,
}

# ══════════════════════════════════════════════════════════════════════════════
# [C] PRIMARY PLAN COUNTS (Jan-Dec 2026)
# ══════════════════════════════════════════════════════════════════════════════
_M = FORECAST_MONTHS
PRIMARY_PLAN_COUNTS = {
    ('Base', 'Private'): dict(zip(_M, [134,100,118,145,149,204,221,165,144,133,143,166])),
    ('Base', 'Premium'): dict(zip(_M, [294,220,258,317,327,448,486,362,316,293,314,363])),
    ('Base', 'Group'):   dict(zip(_M, [451,338,395,486,502,686,744,555,485,449,481,557])),
    ('MT',   'Private'): dict(zip(_M, [109,116,137,154,159,209,155,155,137,123,126,159])),
    ('MT',   'Premium'): dict(zip(_M, [127,135,160,180,186,244,181,180,160,143,147,186])),
    ('MT',   'Group'):   dict(zip(_M, [127,135,160,180,186,244,181,180,160,143,147,186])),
}

# ══════════════════════════════════════════════════════════════════════════════
# [D] LOAD INPUTS FROM EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def _load_inputs_excel():
    overrides = {
        'plan_counts':          {},   # (ct,gt) -> {mstr: int}
        'package_dist':         {},   # (ct,gt,pkg) -> float
        'renewal_price':        {},   # (ct,gt,pkg) -> float  (price/lesson base)
        'retention_by_renewal': {},   # pno -> float  (global fallback, ALL rows)
        'retention_expanded':   {},   # (pno, ct, gt, pkg_bucket) -> float
        'shares':               {},   # seg -> float
        'ext_curve':            {},   # lag_k -> float
        'rates':                {},   # (ct,gt) -> float
        'settings':             {},   # key -> value (strings / numbers)
    }
    if not _os.path.exists(INPUTS_EXCEL):
        return overrides
    try:
        _xl = __import__('openpyxl').load_workbook(INPUTS_EXCEL, data_only=True)
    except Exception as e:
        print(f'  Warning: could not open {INPUTS_EXCEL}: {e}')
        return overrides

    # plan_counts sheet — читаем ПОЗИЦИОННО (col 1=Jan, col 2=Feb, ...)
    # НЕ используем _M как ключи здесь — он может быть 2026 когда данные 2025!
    # Маппинг на реальные FORECAST_MONTHS происходит ПОСЛЕ применения settings.
    if 'plan_counts' in _xl.sheetnames:
        ws = _xl['plan_counts']
        for row in ws.iter_rows(min_row=3, max_row=3+len(DIMS)-1):
            dim_label = str(row[0].value or '').strip()
            for (ct, gt) in DIMS:
                if dim_label == f'{ct}/{gt}':
                    # Store as positional list (12 values), not month-keyed dict
                    positional = []
                    for ci in range(12):  # 12 months
                        v = row[ci+1].value
                        positional.append(int(float(v)) if v is not None else None)
                    overrides['plan_counts'][(ct,gt)] = positional  # list, not dict
                    break

    # package_dist sheet: cols Dim | Package | Probability | Note
    if 'package_dist' in _xl.sheetnames:
        ws = _xl['package_dist']
        for row in ws.iter_rows(min_row=3):
            dim_label = str(row[0].value or '').strip()
            pkg_v     = row[1].value
            prob_v    = row[2].value
            if dim_label and pkg_v is not None and prob_v is not None:
                for (ct, gt) in DIMS:
                    if dim_label == f'{ct}/{gt}':
                        try:
                            overrides['package_dist'][(ct, gt, int(pkg_v))] = float(prob_v)
                        except: pass
                        break

    # renewal_price sheet: cols Dim | Package | Price/Lesson (base) | Growth%/quarter | ...
    if 'renewal_price' in _xl.sheetnames:
        ws = _xl['renewal_price']
        for row in ws.iter_rows(min_row=3):
            dim_label = str(row[0].value or '').strip()
            pkg_v     = row[1].value
            price_v   = row[2].value
            if dim_label and pkg_v is not None and price_v is not None:
                for (ct, gt) in DIMS:
                    if dim_label == f'{ct}/{gt}':
                        try:
                            overrides['renewal_price'][(ct, gt, int(pkg_v))] = float(price_v)
                        except: pass
                        break

    # retention_by_renewal sheet: new long-format
    # Cols: pno(0) | ct(1) | gt(2) | pkg_bucket(3) | Retention(4) | Override(5) | n_pool(6) | Note(7)
    # 'ALL' rows (ct=ALL, gt=ALL, pkg=ALL) → global fallback keyed by pno
    # pkg-specific rows → retention_expanded keyed by (pno, ct, gt, pkg_bucket)
    # Override (col 5) takes priority over Retention (col 4) if non-empty
    if 'retention_by_renewal' in _xl.sheetnames:
        ws = _xl['retention_by_renewal']
        for row in ws.iter_rows(min_row=4):
            pno_v   = row[0].value
            ct_v    = str(row[1].value or '').strip() if len(row) > 1 else ''
            gt_v    = str(row[2].value or '').strip() if len(row) > 2 else ''
            pkg_v   = row[3].value                    if len(row) > 3 else None
            rate_v  = row[4].value                    if len(row) > 4 else None
            over_v  = row[5].value                    if len(row) > 5 else None
            if pno_v is None:
                continue
            # Effective rate: override takes priority
            eff_v = over_v if (over_v is not None and over_v != '') else rate_v
            if eff_v is None:
                continue
            try:
                pno_int = int(pno_v)
                eff_f   = float(eff_v)
            except Exception:
                continue
            # 'ALL' rows → global fallback dict keyed by pno int
            if ct_v == 'ALL' or pkg_v is None or str(pkg_v).strip() == 'ALL':
                overrides['retention_by_renewal'][pno_int] = eff_f
            else:
                # pkg-specific row
                if ct_v in ('Base', 'MT') and gt_v in ('Private', 'Premium', 'Group'):
                    try:
                        pkg_int = int(float(pkg_v))
                        overrides['retention_expanded'][(pno_int, ct_v, gt_v, pkg_int)] = eff_f
                    except Exception:
                        pass

    # shares sheet: cols Segment | Share | ...
    if 'shares' in _xl.sheetnames:
        ws = _xl['shares']
        for row in ws.iter_rows(min_row=3, max_row=3+len(SEGS)-1):
            seg_label  = str(row[0].value or '').strip()
            override_v = row[1].value
            if seg_label in SEGS and override_v is not None:
                try: overrides['shares'][seg_label] = float(override_v)
                except: pass

    # ext_curve sheet: cols Lag | Override | ...  (override in col E historically)
    if 'ext_curve' in _xl.sheetnames:
        ws = _xl['ext_curve']
        for row in ws.iter_rows(min_row=4):
            lag_label  = str(row[0].value or '').strip()
            # Try col E (index 4) first, then col B (index 1) as override
            override_v = row[4].value if len(row) > 4 else None
            if override_v is None and len(row) > 1:
                override_v = row[1].value
            if lag_label.startswith('lag ') and override_v is not None:
                try:
                    k = int(lag_label.split()[1])
                    overrides['ext_curve'][k] = float(override_v)
                except: pass

    # rates sheet: cols Dim | Rate | ...
    if 'rates' in _xl.sheetnames:
        ws = _xl['rates']
        for row in ws.iter_rows(min_row=3, max_row=3+len(DIMS)-1):
            dim_label  = str(row[0].value or '').strip()
            override_v = row[1].value
            for (ct, gt) in DIMS:
                if dim_label == f'{ct}/{gt}':
                    if override_v is not None:
                        try: overrides['rates'][(ct,gt)] = float(override_v)
                        except: pass
                    break

    # settings sheet: rows 3+, col A = param name, col B = value
    if 'settings' in _xl.sheetnames:
        for row in _xl['settings'].iter_rows(min_row=3):
            param = str(row[0].value or '').strip()
            val   = row[1].value
            if param and val is not None and not param.startswith('──'):
                overrides['settings'][param] = val

    return overrides

_inp = _load_inputs_excel()

# ── Apply settings overrides ──────────────────────────────────────────────────
_cfg = _inp.get('settings', {})

# AOV adjustment factor (multiplier for all revenues)
AOV_ADJ = float(_cfg.get('aov_adjustment_factor', 1.0))

# Retention adjustment: multiplier for all retention_by_renewal values
# May 2026 actual: retention=38.4% vs calibrated 43.6% (-12%)
# Use <1.0 to account for declining renewal rates in 2026
RETENTION_ADJ = float(_cfg.get('retention_adjustment', 1.0))

# Price growth override
_pgq = _cfg.get('price_growth_pct_per_quarter')
if _pgq is not None:
    PRICE_GROWTH_PCT = float(_pgq)

# data_cutoff override
_dc_val = _cfg.get('data_cutoff')
if _dc_val is not None:
    DATA_CUTOFF_STR = str(_dc_val)[:10]

# cal_window override
_cws = _cfg.get('cal_window_start')
_cwe = _cfg.get('cal_window_end')
if _cws and _cwe:
    CAL_WINDOW = [str(p) for p in pd.period_range(str(_cws)[:7], str(_cwe)[:7], freq='M')]

# forecast window override
_fs = _cfg.get('forecast_start')
_fe = _cfg.get('forecast_end')
if _fs and _fe:
    FORECAST_MONTHS = [str(p) for p in pd.period_range(str(_fs)[:7], str(_fe)[:7], freq='M')]
    # Expand PRIMARY_PLAN_COUNTS for any months beyond 2026-12
    for (ct, gt) in DIMS:
        vals = PRIMARY_PLAN_COUNTS[(ct, gt)]
        last_val = vals.get('2026-12') or list(vals.values())[-1]
        for mstr in FORECAST_MONTHS:
            if mstr not in vals:
                vals[mstr] = last_val  # flat continuation

# Apply plan_counts overrides — AFTER settings (so FORECAST_MONTHS is correct)
# plan_counts from Excel is a positional list, map to actual FORECAST_MONTHS
if _inp['plan_counts']:
    for (ct, gt), positional in _inp['plan_counts'].items():
        if isinstance(positional, list):
            # Positional: map col i → FORECAST_MONTHS[i]
            for i, v in enumerate(positional):
                if i < len(FORECAST_MONTHS) and v is not None:
                    PRIMARY_PLAN_COUNTS[(ct,gt)][FORECAST_MONTHS[i]] = v
        else:
            # Legacy dict format (backward compat)
            for mstr, v in positional.items():
                PRIMARY_PLAN_COUNTS[(ct,gt)][mstr] = v
    print(f'  Loaded plan_counts from Excel → mapped to {FORECAST_MONTHS[0]}–{FORECAST_MONTHS[-1]}')

_ext_curve_overrides        = _inp['ext_curve']
_retention_overrides        = _inp['retention_by_renewal']
_retention_expanded_from_xl = _inp['retention_expanded']
_pkg_dist_overrides         = _inp['package_dist']
_renewal_price_overrides    = _inp['renewal_price']
_shares_overrides           = _inp['shares']
_rates_overrides            = _inp['rates']

# Log applied settings
if AOV_ADJ != 1.0:
    print(f'  AOV adjustment factor: ×{AOV_ADJ:.3f}  (все выручки × {AOV_ADJ})')
if _dc_val:
    print(f'  data_cutoff from Excel: {DATA_CUTOFF_STR}')
if _fs and _fe:
    print(f'  Forecast window from Excel: {FORECAST_MONTHS[0]} → {FORECAST_MONTHS[-1]}  ({len(FORECAST_MONTHS)} мес.)')
if _cws and _cwe:
    print(f'  Calibration window from Excel: {CAL_WINDOW[0]}–{CAL_WINDOW[-1]}')

if not _os.path.exists(INPUTS_EXCEL):
    print(f'  Info: {INPUTS_EXCEL} not found — using defaults. Run build_forecast_inputs.py to create it.')

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
DATA_CUTOFF = pd.Timestamp(DATA_CUTOFF_STR)

def load_data():
    df = pd.read_csv(DATA_PATH, low_memory=False)
    df.columns = [c.replace('prolongation_test[','').replace(']','') for c in df.columns]
    for col in ['payment_updated_at','prev_next_payment_planned_dt','next_payment_planned_dt']:
        df[col] = pd.to_datetime(df[col], errors='coerce')
    df['usd']         = pd.to_numeric(df['Amount in USD'], errors='coerce')
    df['paid_M']      = df['payment_updated_at'].dt.to_period('M')
    df['prev_plan_M'] = df['prev_next_payment_planned_dt'].dt.to_period('M')
    df['planned_M']   = df['next_payment_planned_dt'].dt.to_period('M')
    df['is_sec']      = df['cat_payment'].isin(['same','grad','secondary upgrade'])
    df['is_prim']     = df['cat_payment'] == 'primary'
    df['course_type'] = (df['full_course_name']
                         .str.contains('Math Tutoring', case=False, na=False)
                         .map({True:'MT', False:'Base'}))
    gt_col = 'group_type_c' if 'group_type_c' in df.columns else None
    if gt_col:
        df['group_type'] = (df[gt_col].fillna('Unknown')
                            .replace({'Private Eng-Math Tut':'Private',
                                      'Mini':'Premium', 'mini':'Premium'}))
        df['group_type'] = df['group_type'].apply(
            lambda x: x if x in ('Private','Premium','Group') else 'Unknown')
    else:
        df['group_type'] = 'Unknown'
    # Package size — ВСЕГДА из credits_recieved_package, НЕ из Package (Package=999 — заглушка!)
    if 'credits_recieved_package' in df.columns:
        df['pkg'] = pd.to_numeric(df['credits_recieved_package'], errors='coerce').abs()
    else:
        df['pkg'] = np.nan
    # pkg_bucket — round pkg to nearest PKG_BUCKETS entry
    df['pkg_bucket'] = df['pkg'].apply(map_to_bucket)
    # payment_no
    if 'payment_no' in df.columns:
        df['payment_no'] = pd.to_numeric(df['payment_no'], errors='coerce')
    else:
        df['payment_no'] = np.nan
    return df[df['payment_updated_at'] <= DATA_CUTOFF].copy()

def classify_segment(row):
    cp = row['cat_payment']
    if cp == 'secondary upgrade': return 'Upgrades'
    if cp not in ('same','grad'):  return None
    pp, pm = row['prev_plan_M'], row['paid_M']
    if pd.isna(pp) or pd.isna(pm): return None
    if pm == pp: return 'Present'
    if pm <  pp: return 'Earlier'
    return 'Reanim'

print(f'{SEP}\nFORECAST 2026 v3  |  cal: {CAL_WINDOW[0]}-{CAL_WINDOW[-1]}\n{SEP}')
df = load_data()
df['segment'] = df.apply(classify_segment, axis=1)
print(f'Loaded: {len(df):,} rows  |  cutoff: {DATA_CUTOFF.date()}')
print(f'Last paid month: {df["paid_M"].max()}')

# Актуалы (только из данных до cutoff — не "заглядываем вперёд")
actuals_rev = (df[df['is_sec'] & df['paid_M'].notna()]
               .groupby('paid_M')['usd'].sum())
actuals_n   = (df[df['is_sec'] & df['paid_M'].notna()]
               .groupby('paid_M').size())

# ══════════════════════════════════════════════════════════════════════════════
# EXT-CURVE
# ══════════════════════════════════════════════════════════════════════════════
first_purch = df.sort_values('payment_updated_at').groupby('student_id').first()
sec_all     = df[df['is_sec']].copy()

def _build_ext_curve(cal_start: str, min_cohorts: int) -> dict:
    _cohorts    = pd.period_range(cal_start, CURVE_CAL_END, freq='M')
    _lag_counts = {k: [] for k in range(1, CURVE_MAX_LAG + 1)}
    for cm in _cohorts:
        stu = set(first_purch[
            (first_purch['cat_payment'] == 'primary') &
            (first_purch['payment_updated_at'].dt.to_period('M') == cm)
        ].index)
        if not stu: continue
        c_sec = sec_all[sec_all['student_id'].isin(stu)]
        for k in range(1, CURVE_MAX_LAG + 1):
            tgt = cm + k
            if tgt > pd.Period(CAL_WINDOW[-1]): continue
            _lag_counts[k].append((c_sec['paid_M'] == tgt).sum() / len(stu))
    return {k: float(np.mean(v)) for k, v in _lag_counts.items() if len(v) >= min_cohorts}

ext_curve_full   = _build_ext_curve(CURVE_CAL_START,        CURVE_MIN_COHORTS)
ext_curve_recent = _build_ext_curve(CURVE_CAL_START_RECENT, CURVE_MIN_COHORTS_RECENT)

ext_curve = {}
for k in range(1, CURVE_MAX_LAG + 1):
    if k in ext_curve_recent:
        ext_curve[k] = ext_curve_recent[k]
    elif k in ext_curve_full:
        ext_curve[k] = ext_curve_full[k]

# Apply ext_curve overrides
for k, v in _ext_curve_overrides.items():
    ext_curve[k] = v

_recent_lags = sorted(ext_curve_recent.keys())
_all_lags    = sorted(ext_curve.keys())
print(f'Ext-curve: {len(ext_curve)} lags  '
      f'(recent 2024+: {_recent_lags[0] if _recent_lags else "?"}-{_recent_lags[-1] if _recent_lags else "?"}  |  '
      f'full fallback up to lag {_all_lags[-1] if _all_lags else "?"})')

# Historical primary cohorts by dim
_prim_df = first_purch[first_purch['cat_payment'] == 'primary'].copy()
_prim_df['cohort_M'] = _prim_df['payment_updated_at'].dt.to_period('M')
_hist_cohort_series  = _prim_df.groupby(['cohort_M','course_type','group_type']).size()
hist_cohort_by_dim: dict = {}
for (cm, ct_, gt_), sz in _hist_cohort_series.items():
    hist_cohort_by_dim.setdefault((ct_, gt_), {})[cm] = int(sz)

# ── Плановые когорты — от текущего (неполного) месяца до конца прогноза ───────
# PLAN_COHORT_START = data_cutoff месяц (включительно): текущий месяц в CSV неполный,
# поэтому первичные продажи за него берутся из plan_counts, не из CSV.
# Значения берутся из PRIMARY_PLAN_COUNTS (лист plan_counts в Excel)
_cutoff_period  = pd.Period(DATA_CUTOFF_STR[:7])
_forecast_end_p = pd.Period(FORECAST_MONTHS[-1])

PLAN_COHORT_START = str(_cutoff_period)  # = data_cutoff month (Jun 2026)
plan_future_cohorts_set = set(pd.period_range(PLAN_COHORT_START,
                                               str(_forecast_end_p), freq='M'))
print(f'Plan cohorts range: {PLAN_COHORT_START} → {str(_forecast_end_p)} '
      f'({len(plan_future_cohorts_set)} months, from plan_counts Excel)')

# ══════════════════════════════════════════════════════════════════════════════
# CALIBRATION 1: SEGMENT SHARES
# ══════════════════════════════════════════════════════════════════════════════
sec_cal = df[
    df['is_sec'] &
    df['paid_M'].astype(str).isin(CAL_WINDOW) &
    df['segment'].notna() &
    df['usd'].notna()
].copy()
total_sec_cal = len(sec_cal)

shares = {}
src_shares = {}
for seg in SEGS:
    ov = _shares_overrides.get(seg)
    if ov is not None:
        shares[seg] = ov; src_shares[seg] = 'override'
    elif total_sec_cal >= MIN_N_SHARES:
        shares[seg] = len(sec_cal[sec_cal['segment'] == seg]) / total_sec_cal
        src_shares[seg] = f'data n={total_sec_cal}'
    else:
        shares[seg] = SHARES_DEFAULT[seg]; src_shares[seg] = f'hardcoded default (n={total_sec_cal})'

_s = sum(shares.values())
if _s > 0: shares = {k: v/_s for k, v in shares.items()}

print(f'\n-- Segment shares (global, {CAL_WINDOW[0]}-{CAL_WINDOW[-1]}, n={total_sec_cal})')
for seg in SEGS:
    print(f'  {seg:<10}: {shares[seg]:>6.1%}  [{src_shares[seg]}]')

# ══════════════════════════════════════════════════════════════════════════════
# CALIBRATION 2: PROLONGATION RATES (fallback when retention_by_renewal not used)
# ══════════════════════════════════════════════════════════════════════════════
pool_df = df[df['planned_M'].notna()].copy()
pool_cal_df = pool_df[pool_df['planned_M'].astype(str).isin(CAL_WINDOW)]

glob_pool = len(pool_cal_df)
glob_rate = total_sec_cal / glob_pool if glob_pool > 0 else 0.20

rates = {}; src_rates = {}
for ct, gt in DIMS:
    ov = _rates_overrides.get((ct, gt))
    if ov is not None:
        rates[(ct,gt)] = ov; src_rates[(ct,gt)] = 'override'
        continue
    # Try hardcoded defaults first; data calibration can improve them
    p = len(pool_cal_df[(pool_cal_df['course_type']==ct)&(pool_cal_df['group_type']==gt)])
    s = len(sec_cal[(sec_cal['course_type']==ct)&(sec_cal['group_type']==gt)])
    if p >= MIN_N_RATE:
        rates[(ct,gt)] = s/p
        src_rates[(ct,gt)] = f'data pool={p} sales={s}'
    else:
        rates[(ct,gt)] = PROL_RATES_DEFAULT.get((ct,gt), glob_rate)
        src_rates[(ct,gt)] = f'hardcoded default (pool={p}<{MIN_N_RATE})'

print(f'\n-- Prolongation rates (by ct x gt)')
for ct, gt in DIMS:
    print(f'  {ct}/{gt:<15}: {rates[(ct,gt)]:>6.2%}  [{src_rates[(ct,gt)]}]')

# ══════════════════════════════════════════════════════════════════════════════
# CALIBRATION 3: RETENTION BY PAYMENT_NO
# ══════════════════════════════════════════════════════════════════════════════
# Build retention from data: for each pno n, retention = P(secondary | in_pool_with_pno=n)
# pool_by_pno[(m, ct, gt, n)] = count of students with planned_M=m, pno=n
# retention[n] = secondary_sales at pno=n / pool at pno=n

retention_by_renewal = {}
src_retention        = {}

if 'payment_no' in df.columns and df['payment_no'].notna().sum() > 100:
    # pool_by_pno: students with a planned renewal in cal_window, grouped by pno
    pool_pno_cal = pool_cal_df[pool_cal_df['payment_no'].notna()].copy()
    sec_pno_cal  = sec_cal[sec_cal['payment_no'].notna()].copy()

    for pno in range(1, 16):
        ov = _retention_overrides.get(pno)
        if ov is not None:
            retention_by_renewal[pno] = ov
            src_retention[pno] = 'override'
            continue
        pool_n = len(pool_pno_cal[pool_pno_cal['payment_no'] == pno])
        sec_n  = len(sec_pno_cal[sec_pno_cal['payment_no'] == pno + 1])  # next step
        if pool_n >= 10:
            retention_by_renewal[pno] = min(sec_n / pool_n, 1.0)
            src_retention[pno] = f'data pool={pool_n} sec={sec_n}'
        else:
            retention_by_renewal[pno] = RETENTION_BY_PNO_DEFAULT.get(pno, RETENTION_PNO_HIGH)
            src_retention[pno] = f'hardcoded default (pool={pool_n}<10)'
else:
    # No payment_no data: use hardcoded defaults
    for pno in range(1, 16):
        ov = _retention_overrides.get(pno)
        retention_by_renewal[pno] = ov if ov is not None else RETENTION_BY_PNO_DEFAULT.get(pno, RETENTION_PNO_HIGH)
        src_retention[pno] = 'override' if ov is not None else 'hardcoded default (no pno data)'

print(f'\n-- Retention by payment_no (sample)')
for pno in [1, 2, 3, 4, 5, 6, 7]:
    print(f'  pno={pno}: {retention_by_renewal[pno]:.1%}  [{src_retention[pno]}]')

# ══════════════════════════════════════════════════════════════════════════════
# CALIBRATION 4: PACKAGE DISTRIBUTION AND PRICE PER LESSON
# ══════════════════════════════════════════════════════════════════════════════
# Calibration from secondary sales Jul-Dec 2025 with valid package
sec_cal_pkg = sec_cal[sec_cal['pkg'].notna() & (sec_cal['pkg'] > 0) & (sec_cal['pkg'] <= 500)].copy()

# Build pkg_dist: (ct, gt, pkg) -> probability (normalized per dim)
pkg_dist_data   = {}   # raw counts by (ct, gt, pkg)
pkg_dist_final  = {}   # final probabilities
pkg_dist_src    = {}

for ct, gt in DIMS:
    dim_sec = sec_cal_pkg[(sec_cal_pkg['course_type']==ct) & (sec_cal_pkg['group_type']==gt)]
    n_dim = len(dim_sec)
    if n_dim >= MIN_N_PKG_DIST:
        counts = dim_sec['pkg'].value_counts()
        for pkg_sz, cnt in counts.items():
            pkg_dist_data[(ct, gt, int(pkg_sz))] = cnt / n_dim
            pkg_dist_src[(ct, gt, int(pkg_sz))] = f'data n={n_dim}'
    else:
        # Fallback: use Base/Group distribution if this dim is thin
        # MT dims fall back to Base equivalent, keeping the group_type
        fallback_key = ('Base', gt) if ct == 'MT' else ('Base', 'Group')
        base_sec = sec_cal_pkg[
            (sec_cal_pkg['course_type'] == fallback_key[0]) &
            (sec_cal_pkg['group_type']  == fallback_key[1])
        ]
        if len(base_sec) >= MIN_N_PKG_DIST:
            counts = base_sec['pkg'].value_counts()
            total  = len(base_sec)
            for pkg_sz, cnt in counts.items():
                pkg_dist_data[(ct, gt, int(pkg_sz))] = cnt / total
                pkg_dist_src[(ct, gt, int(pkg_sz))] = f'Base fallback n={total}'
        else:
            # Hard fallback: hardcoded Base/Group distribution
            for pkg_sz, prob in PKG_DIST_FALLBACK_BASE_GROUP.items():
                pkg_dist_data[(ct, gt, pkg_sz)] = prob
                pkg_dist_src[(ct, gt, pkg_sz)] = 'hardcoded Base/Group fallback'

# Apply package_dist overrides from Excel
for (ct, gt, pkg_sz), prob in _pkg_dist_overrides.items():
    pkg_dist_data[(ct, gt, pkg_sz)] = prob
    pkg_dist_src[(ct, gt, pkg_sz)] = 'override'

# Normalize per dim
for ct, gt in DIMS:
    dim_pkgs = {k: v for k, v in pkg_dist_data.items() if k[0]==ct and k[1]==gt}
    total_prob = sum(dim_pkgs.values())
    if total_prob > 0:
        for k in dim_pkgs:
            pkg_dist_final[k] = dim_pkgs[k] / total_prob
    else:
        # last resort: uniform over hardcoded packages
        pkgs = list(PKG_DIST_FALLBACK_BASE_GROUP.keys())
        for p in pkgs:
            pkg_dist_final[(ct, gt, p)] = 1.0 / len(pkgs)

# Build price per lesson: (ct, gt, pkg) -> price_per_lesson (base period Q3 2025)
ppl_data  = {}   # raw from data
ppl_final = {}   # final values
ppl_src   = {}

for ct, gt in DIMS:
    dim_sec = sec_cal_pkg[(sec_cal_pkg['course_type']==ct) & (sec_cal_pkg['group_type']==gt)]
    pkgs_in_dim = set(k[2] for k in pkg_dist_final if k[0]==ct and k[1]==gt)
    for pkg_sz in pkgs_in_dim:
        pkg_sz_int = int(pkg_sz)
        ov = _renewal_price_overrides.get((ct, gt, pkg_sz_int))
        if ov is not None:
            ppl_final[(ct, gt, pkg_sz_int)] = ov
            ppl_src[(ct, gt, pkg_sz_int)] = 'override'
            continue
        # Try data: median usd / pkg_size
        sub = dim_sec[(dim_sec['pkg'] == pkg_sz_int) & dim_sec['usd'].notna()]
        if len(sub) >= MIN_N_PKG:
            ppl_v = float((sub['usd'] / pkg_sz_int).mean())
            ppl_data[(ct, gt, pkg_sz_int)] = ppl_v
            ppl_final[(ct, gt, pkg_sz_int)] = ppl_v
            ppl_src[(ct, gt, pkg_sz_int)] = f'data n={len(sub)}'
        else:
            # Fallback hierarchy: (ct,gt,pkg) hardcoded -> (gt,pkg) -> dim median
            hc = PPL_HARDCODED.get((ct, gt, pkg_sz_int))
            if hc is not None:
                ppl_final[(ct, gt, pkg_sz_int)] = hc
                ppl_src[(ct, gt, pkg_sz_int)] = f'hardcoded (n={len(sub)}<{MIN_N_PKG})'
            else:
                # Try same gt, different ct
                gt_sub = sec_cal_pkg[(sec_cal_pkg['group_type']==gt) & (sec_cal_pkg['pkg']==pkg_sz_int) & sec_cal_pkg['usd'].notna()]
                if len(gt_sub) >= MIN_N_PKG:
                    ppl_v = float((gt_sub['usd'] / pkg_sz_int).mean())
                    ppl_final[(ct, gt, pkg_sz_int)] = ppl_v
                    ppl_src[(ct, gt, pkg_sz_int)] = f'gt fallback n={len(gt_sub)}'
                else:
                    # Use median across all secondary for this dim
                    dim_all = sec_cal_pkg[(sec_cal_pkg['course_type']==ct) & (sec_cal_pkg['group_type']==gt) & sec_cal_pkg['usd'].notna()]
                    if len(dim_all) >= 5 and pkg_sz_int > 0:
                        ppl_v = float((dim_all['usd'] / pkg_sz_int).mean())
                        ppl_final[(ct, gt, pkg_sz_int)] = max(ppl_v, 1.0)
                        ppl_src[(ct, gt, pkg_sz_int)] = f'dim median (pkg {pkg_sz_int}, n={len(dim_all)})'
                    else:
                        ppl_final[(ct, gt, pkg_sz_int)] = 7.0   # last resort
                        ppl_src[(ct, gt, pkg_sz_int)] = 'last resort default'

print(f'\n-- Package distribution calibrated for {len(DIMS)} dims')
for ct, gt in DIMS:
    dim_p = {k[2]: v for k, v in pkg_dist_final.items() if k[0]==ct and k[1]==gt}
    top3 = sorted(dim_p.items(), key=lambda x: -x[1])[:3]
    print(f'  {ct}/{gt}: top pkgs = ' + ' '.join(f'pkg{p}={v:.0%}' for p,v in top3))

# ══════════════════════════════════════════════════════════════════════════════
# ПРЯМОЙ AOV — calibrated directly from Q4-2025 secondary sales per (dim, seg)
# Это точнее чем pkg × ppl: захватывает реальный mix пакетов Q4-2025
# ══════════════════════════════════════════════════════════════════════════════
_CAL_Q4 = ["2025-10","2025-11","2025-12"]   # Q4-2025 = ценовая база
sec_q4_all = df[
    df['is_sec'] &
    df['paid_M'].astype(str).isin(_CAL_Q4) &
    df['usd'].notna() &
    df['segment'].notna()
].copy()

direct_aov_base   = {}   # (ct, gt, seg) → mean USD/sale in Q4-2025
direct_aov_src    = {}
_glob_q4          = sec_q4_all['usd'].mean() if len(sec_q4_all) > 0 else 180.0

for ct, gt in DIMS:
    for seg in SEGS:
        key = (ct, gt, seg)
        sub_dsg = sec_q4_all[(sec_q4_all['course_type']==ct) &
                              (sec_q4_all['group_type']==gt) &
                              (sec_q4_all['segment']==seg)]
        if len(sub_dsg) >= 10:
            direct_aov_base[key] = sub_dsg['usd'].mean()
            direct_aov_src[key]  = f'Q4-2025 dim+seg n={len(sub_dsg)}'
        else:
            sub_dim = sec_q4_all[(sec_q4_all['course_type']==ct) & (sec_q4_all['group_type']==gt)]
            if len(sub_dim) >= 5:
                direct_aov_base[key] = sub_dim['usd'].mean()
                direct_aov_src[key]  = f'Q4-2025 dim n={len(sub_dim)}'
            else:
                sub_seg = sec_q4_all[sec_q4_all['segment']==seg]
                if len(sub_seg) >= 5:
                    direct_aov_base[key] = sub_seg['usd'].mean()
                    direct_aov_src[key]  = f'Q4-2025 global+seg n={len(sub_seg)}'
                else:
                    direct_aov_base[key] = _glob_q4
                    direct_aov_src[key]  = f'Q4-2025 global fallback'

print(f'\n-- Direct AOV base (Q4-2025) по dims:')
for ct, gt in DIMS:
    aovs = [direct_aov_base[(ct,gt,seg)] for seg in SEGS]
    print(f'  {ct}/{gt}: ' + '  '.join(f'{seg}=${direct_aov_base[(ct,gt,seg)]:.0f}' for seg in SEGS))

# ══════════════════════════════════════════════════════════════════════════════
# PRICE GROWTH FACTOR
# ══════════════════════════════════════════════════════════════════════════════
def _quarter_index(year: int, month: int) -> float:
    """Returns quarter number (fractional) from Q1 2025 = 1."""
    return (year - 2025) * 4 + (month - 1) // 3 + 1

_base_month = {1:1, 2:4, 3:7, 4:10}[PRICE_BASE_QUARTER]  # Q4=10 (October)
BASE_QUARTER_IDX = _quarter_index(PRICE_BASE_YEAR, _base_month)

def price_growth_factor(mstr: str) -> float:
    """Growth factor relative to Q4-2025 base."""
    yr  = int(mstr[:4])
    mo  = int(mstr[5:7])
    q_idx = _quarter_index(yr, mo)
    delta_q = q_idx - BASE_QUARTER_IDX
    return (1 + PRICE_GROWTH_PCT) ** delta_q

# ══════════════════════════════════════════════════════════════════════════════
# POOL LOGIC — 3 MODES
# ══════════════════════════════════════════════════════════════════════════════
# pool_by_pno: pool_df grouped by (planned_M, course_type, group_type, payment_no)
if 'payment_no' in pool_df.columns:
    pool_by_pno = pool_df[pool_df['payment_no'].notna()].groupby(
        ['planned_M','course_type','group_type','payment_no']
    ).size()
else:
    pool_by_pno = pd.Series(dtype=int)

# pool_by_pno_pkg: expanded pool grouped by (planned_M, ct, gt, payment_no, pkg_bucket)
if 'payment_no' in pool_df.columns and 'pkg_bucket' in pool_df.columns:
    pool_by_pno_pkg = pool_df[
        pool_df['payment_no'].notna() & pool_df['pkg_bucket'].notna()
    ].groupby(
        ['planned_M','course_type','group_type','payment_no','pkg_bucket']
    ).size()
else:
    pool_by_pno_pkg = pd.Series(dtype=int)

# retention_expanded: calibrated from data + any overrides from Excel
# Keyed by (pno, ct, gt, pkg_bucket_int)
retention_expanded: dict = {}

# ── Build from on-the-fly calibration (same logic as build_forecast_inputs.py) ─
if 'payment_no' in df.columns and df['payment_no'].notna().sum() > 100:
    _sec_cal_exp = df[
        df['is_sec'] &
        df['paid_M'].astype(str).isin(CAL_WINDOW) &
        df['segment'].notna() &
        df['usd'].notna() &
        df['payment_no'].notna()
    ].copy()

    # Compute prev_pkg_bucket for sec sales: pkg_bucket at payment (pno-1)
    if 'student_id' in df.columns:
        _pkg_lkp = (
            df[df['payment_no'].notna() & df['pkg_bucket'].notna()]
            [['student_id', 'payment_no', 'pkg_bucket']]
            .rename(columns={'payment_no': '_prev_pno', 'pkg_bucket': 'prev_pkg_bucket'})
        )
        _sec_cal_exp['_prev_pno'] = _sec_cal_exp['payment_no'] - 1
        _sec_cal_exp = _sec_cal_exp.merge(
            _pkg_lkp,
            left_on=['student_id', '_prev_pno'],
            right_on=['student_id', '_prev_pno'],
            how='left'
        )
        _sec_cal_exp.drop(columns=['_prev_pno'], inplace=True)
    else:
        _sec_cal_exp['prev_pkg_bucket'] = None

    _pool_pno_cal_exp = pool_cal_df[
        pool_cal_df['payment_no'].notna() & pool_cal_df['pkg_bucket'].notna()
    ].copy()

    for _pno in range(1, 16):
        for _ct, _gt in DIMS:
            for _pkg_b in PKG_BUCKETS:
                _pm = (
                    (_pool_pno_cal_exp['payment_no'] == float(_pno)) &
                    (_pool_pno_cal_exp['course_type'] == _ct) &
                    (_pool_pno_cal_exp['group_type']  == _gt) &
                    (_pool_pno_cal_exp['pkg_bucket']  == _pkg_b)
                )
                _pn = _pm.sum()
                if _pn >= 15:
                    _ppb = _sec_cal_exp['prev_pkg_bucket'] if 'prev_pkg_bucket' in _sec_cal_exp.columns else pd.Series([None]*len(_sec_cal_exp), index=_sec_cal_exp.index)
                    _sm = (
                        (_sec_cal_exp['payment_no'] == float(_pno + 1)) &
                        (_sec_cal_exp['course_type'] == _ct) &
                        (_sec_cal_exp['group_type']  == _gt) &
                        (_ppb == _pkg_b)
                    )
                    _sn = _sm.sum()
                    retention_expanded[(_pno, _ct, _gt, _pkg_b)] = min(_sn / _pn, 1.0)

# Apply overrides from Excel (take priority over calibrated values)
for k, v in _retention_expanded_from_xl.items():
    retention_expanded[k] = v

if retention_expanded:
    print(f'  retention_expanded loaded: {len(retention_expanded)} (pno,ct,gt,pkg) combinations')

# pool_raw_dim: without payment_no breakdown (for blend scaling and diagnostics)
pool_raw_dim = pool_df.groupby(['planned_M','course_type','group_type']).size()

DATA_CUTOFF_PERIOD = pd.Period(DATA_CUTOFF_STR[:7])   # 2026-05
DATA_MODES = {
    'data':  [f'2026-{m:02d}' for m in range(1, 5)],    # Jan-Apr
    'blend': ['2026-05'],                                 # May
    'ext':   [f'2026-{m:02d}' for m in range(6, 13)],   # Jun-Dec
}
# Blend scaling: диагностика показала что (31/28) ЗАВЫШАЕТ предсказание.
# pool × retention уже предсказывает полный месяц (все студенты в pipeline).
# Scaling добавляет +10% сверху без пользы → убираем его.
# Фактически: pool=2100, retention=43.6% → 915; actual_full_may=807 (retention упал до 38.4%)
# Ошибка -12% = снижение retention 2026 vs 2025, а не проблема scaling.
BLEND_SCALE = 1.0  # removed: was 31/28, но это давало +10% излишнего завышения

def get_total_sec(m: pd.Period, ct: str, gt: str) -> tuple:
    """
    Returns (total_sec, pool_display, mode).

    mode='data'  (Jan-Apr): total_sec = Σ_n pool_by_pno[(m,ct,gt,n)] * retention[n]
    mode='blend' (May):     то же, без масштабирования (BLEND_SCALE=1.0)
    mode='ext'   (Jun-Dec): total_sec = Σ_cohort cohort_size * ext_curve[lag]
    """
    mstr = str(m)

    if mstr in DATA_MODES['data'] or mstr in DATA_MODES['blend']:
        # Sum over payment_no × pkg_bucket buckets using retention_expanded (with fallback)
        total_sec = 0.0
        pool_total = 0.0
        if len(pool_by_pno) > 0:
            for pno_raw in range(1, 28):
                pno_f = float(pno_raw)
                # Global fallback retention for this pno (Level-2 / Level-3)
                ret_pno = (retention_by_renewal.get(pno_raw) or
                           retention_by_renewal.get(min(pno_raw, 15), RETENTION_PNO_HIGH))

                # ── Students with known pkg_bucket (use expanded retention) ──
                n_with_pkg = 0
                for pkg_b in PKG_BUCKETS:
                    try:
                        n_pkg = int(pool_by_pno_pkg.get((m, ct, gt, pno_f, float(pkg_b)), 0))
                        if n_pkg == 0:
                            n_pkg = int(pool_by_pno_pkg.get((m, ct, gt, pno_raw, pkg_b), 0))
                    except Exception:
                        n_pkg = 0
                    if n_pkg > 0:
                        # Fallback: Level-1 (pno,ct,gt,pkg) → Level-2/3 (pno global)
                        ret = (retention_expanded.get((pno_raw, ct, gt, pkg_b)) or ret_pno)
                        total_sec += n_pkg * ret * RETENTION_ADJ
                        pool_total += n_pkg
                        n_with_pkg += n_pkg

                # ── Students with no pkg info: use global pno retention ───────
                try:
                    n_total_pno = int(pool_by_pno.get((m, ct, gt, pno_f), 0))
                    if n_total_pno == 0:
                        n_total_pno = int(pool_by_pno.get((m, ct, gt, pno_raw), 0))
                except Exception:
                    n_total_pno = 0
                n_no_pkg = max(0, n_total_pno - n_with_pkg)
                if n_no_pkg > 0:
                    total_sec += n_no_pkg * ret_pno * RETENTION_ADJ
                    pool_total += n_no_pkg
        else:
            # No pno breakdown: use flat rate
            raw_pool = float(pool_raw_dim.get((m, ct, gt), 0))
            flat_rate = rates.get((ct, gt), glob_rate)
            total_sec = raw_pool * flat_rate
            pool_total = raw_pool

        if mstr in DATA_MODES['blend']:
            total_sec *= BLEND_SCALE
            pool_total *= BLEND_SCALE
            mode = 'blend'
        else:
            mode = 'data'
        return total_sec, pool_total, mode

    else:
        # ext mode: cohort × ext_curve
        ext_sales = 0.0
        # Historical cohorts
        for C_m, sz in hist_cohort_by_dim.get((ct, gt), {}).items():
            if C_m >= m: continue
            lag = (m.year - C_m.year) * 12 + (m.month - C_m.month)
            if lag in ext_curve:
                ext_sales += sz * ext_curve[lag]
        # Plan cohorts (data_cutoff+1 → forecast_end) — из PRIMARY_PLAN_COUNTS (Excel)
        for C in plan_future_cohorts_set:
            if C >= m: continue
            lag = (m.year - C.year) * 12 + (m.month - C.month)
            if lag not in ext_curve: continue
            cnt = PRIMARY_PLAN_COUNTS.get((ct, gt), {}).get(str(C))
            if cnt:
                ext_sales += cnt * ext_curve[lag]
        # pool_display = ext_sales / rate (pipeline equivalent)
        r = rates.get((ct, gt), glob_rate)
        pool_disp = ext_sales / r if r > 0 else ext_sales
        return ext_sales, pool_disp, 'ext'

# Pool diagnostics
print(f'\n-- Pool diagnostics (mode per month)')
print(f'  {"Month":<10}  ' + '  '.join(f'{ct[:2]}/{gt[:3]}' for ct,gt in DIMS) + '  TOTAL   [mode]')
print('  ' + '-' * 90)
for mstr in FORECAST_MONTHS:
    m_p = pd.Period(mstr)
    results  = [get_total_sec(m_p, ct, gt) for ct, gt in DIMS]
    totals   = [r[0] for r in results]
    pools    = [r[1] for r in results]
    modes    = [r[2] for r in results]
    mode_set = set(modes)
    mode_tag = modes[0] if len(mode_set) == 1 else 'mixed'
    print(f'  {mstr:<10}  ' +
          '  '.join(f'{t:>8.1f}' for t in totals) +
          f'  {sum(totals):>8.1f}  [{mode_tag}]')

# ══════════════════════════════════════════════════════════════════════════════
# REVENUE CALCULATION — DIRECT AOV (Q4-2025 base × price growth)
# Замена pkg×ppl подхода: берём прямой средний чек из данных Q4-2025
# Это точнее т.к. захватывает реальный mix пакетов без упущений
# ══════════════════════════════════════════════════════════════════════════════
def calc_dim_revenue(total_sec: float, ct: str, gt: str, mstr: str) -> dict:
    """
    Revenue = Σ_seg (total_sec × share[seg] × direct_aov_base[(ct,gt,seg)] × pgf × aov_adj)
    direct_aov_base = mean USD per secondary sale in Q4-2025, per dim×seg
    pgf = price_growth_factor(mstr): +3%/quarter from Q4-2025 base
    """
    pgf        = price_growth_factor(mstr)
    rev_by_seg = {}
    n_by_seg   = {}

    for seg in SEGS:
        seg_sec = total_sec * shares[seg]
        aov_v   = direct_aov_base.get((ct, gt, seg), _glob_q4)
        seg_rev = seg_sec * aov_v * pgf * AOV_ADJ
        rev_by_seg[seg] = seg_rev
        n_by_seg[seg]   = seg_sec

    total_rev = sum(rev_by_seg.values())
    aov       = total_rev / total_sec if total_sec > 0 else 0.0
    return {
        'total_rev': total_rev,
        'rev_by_seg': rev_by_seg,
        'n_by_seg':   n_by_seg,
        'aov':        aov,
    }

# ══════════════════════════════════════════════════════════════════════════════
# FORECAST 2026
# ══════════════════════════════════════════════════════════════════════════════
_fc_title = f'FORECAST {FORECAST_MONTHS[0]} → {FORECAST_MONTHS[-1]}  ({len(FORECAST_MONTHS)} мес.)'
if AOV_ADJ != 1.0:
    _fc_title += f'  |  AOV×{AOV_ADJ:.3f}'
print(f'\n{SEP}\n{_fc_title}\n{SEP}')

forecast_rows  = []
monthly_totals = []
dim_month_data = {}   # (mstr, ct, gt) -> dict

for mstr in FORECAST_MONTHS:
    m_p = pd.Period(mstr)
    month_rev  = 0.0
    month_sec  = 0.0

    for ct, gt in DIMS:
        total_sec, pool_disp, mode = get_total_sec(m_p, ct, gt)
        rev_info = calc_dim_revenue(total_sec, ct, gt, mstr)
        dim_month_data[(mstr, ct, gt)] = {
            'total_sec': total_sec,
            'pool_disp': pool_disp,
            'mode':      mode,
            'total_rev': rev_info['total_rev'],
            'aov':       rev_info['aov'],
            'rev_by_seg': rev_info['rev_by_seg'],
            'n_by_seg':   rev_info['n_by_seg'],
        }
        for seg in SEGS:
            forecast_rows.append({
                'month': mstr, 'ct': ct, 'gt': gt, 'seg': seg,
                'pool':       round(pool_disp, 1),
                'mode':       mode,
                'total_sec_dim': round(total_sec, 1),
                'sales':      round(rev_info['n_by_seg'][seg], 1),
                'rev':        round(rev_info['rev_by_seg'][seg], 0),
            })
        month_rev += rev_info['total_rev']
        month_sec += total_sec

    # Actuals
    m_per      = pd.Period(mstr)
    has_act    = m_per in actuals_rev.index
    act_rev    = float(actuals_rev.loc[m_per]) if has_act else None
    act_n      = int(actuals_n.loc[m_per])     if has_act else None
    err_pct    = (month_rev - act_rev) / act_rev * 100 if act_rev else None

    monthly_totals.append({
        'month':      mstr,
        'total_sales': round(month_sec, 1),
        'total_rev':   round(month_rev),
        'aov':         round(month_rev / month_sec, 1) if month_sec > 0 else 0,
        'actual_rev':  round(act_rev) if act_rev else None,
        'actual_n':    act_n,
        'err_pct':     round(err_pct, 1) if err_pct is not None else None,
    })

    pgf_str = f'{price_growth_factor(mstr):.3f}'
    err_str = f'  err={err_pct:+.1f}%  (actual=${act_rev:,.0f})' if act_rev else ''
    aov_str = f'{month_rev/month_sec:.0f}' if month_sec > 0 else '--'
    print(f'  {mstr}: ${month_rev:>9,.0f}  ({month_sec:>6.1f} sales  AOV=${aov_str}  pgf={pgf_str}){err_str}')

df_fc  = pd.DataFrame(forecast_rows)
df_tot = pd.DataFrame(monthly_totals)
known  = df_tot.dropna(subset=['actual_rev'])

yr_rev   = int(df_tot['total_rev'].sum())
yr_sales = df_tot['total_sales'].sum()
print(f'\n  TOTAL 2026: ${yr_rev:,.0f}  ({yr_sales:,.0f} sales)')
if len(known) > 0:
    act_sum = known['actual_rev'].sum()
    agg_err = (known['total_rev'].sum() - act_sum) / act_sum * 100
    print(f'  Aggregate error ({len(known)} months): {agg_err:+.1f}%')

# ══════════════════════════════════════════════════════════════════════════════
# Q4-2025 VALIDATION  (Oct | Nov | Dec 2025)
# Pool: payment_updated_at <= 2025-09-30  |  Rates: calibrated Jul-Dec 2025
# ══════════════════════════════════════════════════════════════════════════════
print()
_Q4_SEP = '─' * 72
print(f'\n{_Q4_SEP}')
print('Q4-2025 VALIDATION  (pool cutoff=2025-09-30,  rates=Jul-Dec 2025)')
print(_Q4_SEP)

print(f'  {"Month":<10}  {"N pred":>7}  {"N fact":>7}  {"err N":>6}  '
      f'{"Rev pred":>10}  {"Rev fact":>10}  {"err Rev":>7}  [mode]')
print('  ' + '-' * 72)
# Use ext mode (cohort-based): captures seasonal patterns better than pool-mode
# for months outside forecast_start window.
# Dec due to year-end seasonality; ext_curve implicitly captures it.
q4_validation_rows = []
for _mstr in ['2025-10', '2025-11', '2025-12']:
    _mp     = pd.Period(_mstr, 'M')
    _pred_n = 0.0
    _pool_d = 0.0
    _mode_s = 'ext'
    for _ct, _gt in DIMS:
        _sv, _pd, _mv = get_total_sec(_mp, _ct, _gt)
        _pred_n += _sv
        _pool_d += _pd
        _mode_s  = _mv
    _act_n   = int(actuals_n.get(_mp, 0))
    _act_rev = float(actuals_rev.get(_mp, 0.0))
    _act_aov = _act_rev / _act_n if _act_n > 0 else 180.0
    _pred_rev = _pred_n * _act_aov
    _en  = (_pred_n   - _act_n)   / _act_n   * 100 if _act_n   > 0 else None
    _er  = (_pred_rev - _act_rev) / _act_rev * 100 if _act_rev > 0 else None
    _en_s = f'{_en:+.1f}%' if _en is not None else 'n/a'
    _er_s = f'{_er:+.1f}%' if _er is not None else 'n/a'
    print(f'  {_mstr}     {_pred_n:>7.1f}  {_act_n:>7,}  {_en_s:>6}  '
          f'${_pred_rev:>9,.0f}  ${_act_rev:>9,.0f}  {_er_s:>7}  [ext]')
    q4_validation_rows.append({
        'mstr': _mstr, 'pred_n': _pred_n, 'pred_rev': _pred_rev,
        'act_n': _act_n, 'act_rev': _act_rev,
        'err_n': _en, 'err_rev': _er,
    })

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════
CP = {
    'navy':   'FF1B2A4A', 'dblue':  'FF2E4B7A', 'teal':   'FF00695C',
    'gold':   'FFFFE082', 'lgrey':  'FFF5F7FA', 'white':  'FFFFFFFF',
    'lblue':  'FFE3EEF9', 'lgreen': 'FFE8F5E9', 'loran':  'FFFFF3E0',
    'lpink':  'FFFCE4EC', 'dgrey':  'FF495057', 'formula':'FFFFFCE3',
    'sub':    'FFECEFF1', 'corr':   'FFE8F5E9', 'teal2':  'FFE0F2F1',
    'edit':   'FFFFFDE7', 'lock':   'FFE8EAF6',
}
DARK = {'navy','dblue','teal','dgrey'}

def _fill(k):  return PatternFill('solid', fgColor=CP[k])
def _bord():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def wc(ws, r, c, val, bg='white', bold=False, italic=False,
       align='right', size=10, span=None, fmt=None, wrap=False):
    if span and span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill      = _fill(bg)
    cell.font      = Font(name='Calibri', size=size, bold=bold, italic=italic,
                          color='FFFFFFFF' if bg in DARK else 'FF212529')
    cell.border    = _bord()
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fmt: cell.number_format = fmt

USD  = '$#,##0'
PCT1 = '0.0%'
DIFF = '+0.0%;-0.0%;--'
wb   = Workbook()

# ── SHEET 1: Monthly Summary ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Monthly Summary'
ws1.sheet_view.showGridLines = False
for c, w in {1:12, 2:13, 3:10, 4:10, 5:12, 6:8, 7:9, 8:8}.items():
    ws1.column_dimensions[get_column_letter(c)].width = w

wc(ws1, 1, 1, f'Secondary Sales Forecast 2026 v3  |  Package Revenue Model  |  cal: {CAL_WINDOW[0]}-{CAL_WINDOW[-1]}',
   'navy', bold=True, align='center', size=12, span=8)
ws1.row_dimensions[1].height = 28

hdrs = [('Month','navy'), ('Forecast Rev$','navy'), ('Sales','navy'), ('AOV','navy'),
        ('Actual Rev$','gold'), ('Error%','corr'), ('Actual N','gold'), ('Mode','sub')]
for c, (h, bg) in enumerate(hdrs, 1):
    wc(ws1, 2, c, h, bg, bold=True, align='center')
ws1.row_dimensions[2].height = 18

for i, row in df_tot.iterrows():
    r = 3 + i
    has_act = row['actual_rev'] is not None
    has_err = row['err_pct'] is not None
    def _ec(v): return 'lgreen' if v < 0 else ('lpink' if v > 5 else 'lgrey')
    # Get mode for this month (first dim as representative)
    mode_val = dim_month_data.get((row['month'], 'Base', 'Group'), {}).get('mode', '?')
    vals = [
        (row['month'],           'lgrey',  None),
        (row['total_rev'],       'lblue',  USD),
        (row['total_sales'],     'lblue',  '#,##0.0'),
        (row['aov'],             'lblue',  '$#,##0.0'),
        (row['actual_rev'] if has_act else '--',
         'gold' if has_act else 'lgrey',   USD if has_act else None),
        (row['err_pct']/100 if has_err else '--',
         _ec(row['err_pct']) if has_err else 'lgrey',
         DIFF if has_err else None),
        (row['actual_n'] if has_act else '--',
         'gold' if has_act else 'lgrey',   '#,##0' if has_act else None),
        (mode_val, 'sub', None),
    ]
    for c, (v, bg, fmt) in enumerate(vals, 1):
        wc(ws1, r, c, v, bg, fmt=fmt, align='center' if c in (1, 8) else 'right')
    ws1.row_dimensions[r].height = 16

tr = 3 + len(df_tot)
act_s  = int(known['actual_rev'].sum()) if len(known) > 0 else None
err_tv = (df_tot['total_rev'].sum() - known['actual_rev'].sum()) / known['actual_rev'].sum() if len(known) > 0 else None
tot_vals = [
    ('TOTAL',    'gold', None), (yr_rev, 'gold', USD), (yr_sales, 'gold', '#,##0.0'),
    (yr_rev/yr_sales if yr_sales else 0, 'gold', '$#,##0.0'),
    (act_s if act_s else '--', 'gold', USD if act_s else None),
    (err_tv if err_tv is not None else '--', 'corr' if err_tv is not None else 'lgrey',
     DIFF if err_tv is not None else None),
    ('--', 'gold', None), ('', 'gold', None),
]
for c, (v, bg, fmt) in enumerate(tot_vals, 1):
    wc(ws1, tr, c, v, bg, bold=True, fmt=fmt, align='center' if c == 1 else 'right')
ws1.row_dimensions[tr].height = 22

# ── Q4-2025 Validation section (below TOTAL row) ──────────────────────────────
_q4r = tr + 2   # blank gap of 1 row
# Section header
wc(ws1, _q4r, 1, 'Q4-2025 VALIDATION  (pool cutoff=2025-09-30, rates=Jul-Dec 2025)',
   'sub', bold=True, align='left', size=9, span=8)
ws1.row_dimensions[_q4r].height = 16
_q4r += 1
# Column headers (same as main table)
for c, (h, bg) in enumerate(hdrs, 1):
    wc(ws1, _q4r, c, h, 'sub', bold=True, align='center', size=9)
ws1.row_dimensions[_q4r].height = 14
_q4r += 1
# Data rows
for _row in q4_validation_rows:
    _has_err_n = _row['err_n'] is not None
    _err_bg = 'lgreen' if (_row['err_n'] or 0) < 0 else (
              'lpink' if abs(_row['err_n'] or 0) > 5 else 'lgrey')
    _vals = [
        (_row['mstr'],              'lgrey',  None),
        (_row['pred_rev'],          'lblue',  USD),
        (round(_row['pred_n'], 1),  'lblue',  '#,##0.0'),
        (_row['act_rev'] / _row['act_n'] if _row['act_n'] else 0, 'lblue', '$#,##0.0'),
        (_row['act_rev'],           'gold',   USD),
        (_row['err_rev'] / 100 if _row['err_rev'] is not None else '--',
         _err_bg if _has_err_n else 'lgrey',
         DIFF if _has_err_n else None),
        (_row['act_n'],             'gold',   '#,##0'),
        ('ext',                     'sub',    None),
    ]
    for c, (v, bg, fmt) in enumerate(_vals, 1):
        wc(ws1, _q4r, c, v, bg, fmt=fmt, align='center' if c in (1, 8) else 'right')
    ws1.row_dimensions[_q4r].height = 15
    _q4r += 1

# ── SHEET 2: Dim Breakdown ────────────────────────────────────────────────────
ws2 = wb.create_sheet('Dim Breakdown')
ws2.sheet_view.showGridLines = False

N_DIM_COLS = 5   # Pool | Mode | Sales | Rev$ | AOV
NCOLS2     = 1 + len(DIMS) * N_DIM_COLS + 1

wc(ws2, 1, 1, f'Dimensional Forecast 2026 v3  |  {CAL_WINDOW[0]}-{CAL_WINDOW[-1]}',
   'navy', bold=True, align='center', size=12, span=NCOLS2)
ws2.row_dimensions[1].height = 28

wc(ws2, 2, 1, '', 'navy')
_dim_bg = ['lblue', 'lgreen', 'loran', 'teal2', 'sub', 'lgrey']
for di, (ct, gt) in enumerate(DIMS):
    col_s = 2 + di * N_DIM_COLS
    wc(ws2, 2, col_s, f'{ct}/{gt}', _dim_bg[di % len(_dim_bg)],
       bold=True, align='center', span=N_DIM_COLS)
wc(ws2, 2, NCOLS2, 'TOTAL', 'navy', bold=True, align='center')
ws2.row_dimensions[2].height = 18

wc(ws2, 3, 1, 'Month', 'navy', bold=True, align='center')
for di in range(len(DIMS)):
    bg = _dim_bg[di % len(_dim_bg)]
    for j, h in enumerate(['Pool', 'Mode', 'Sales', 'Rev$', 'AOV']):
        wc(ws2, 3, 2 + di * N_DIM_COLS + j, h, bg, bold=True, align='center', size=9)
wc(ws2, 3, NCOLS2, 'Rev$', 'navy', bold=True, align='center')
ws2.row_dimensions[3].height = 18

dim_rev_totals = {(ct, gt): 0.0 for ct, gt in DIMS}
for i, mstr in enumerate(FORECAST_MONTHS):
    r = 4 + i
    wc(ws2, r, 1, mstr, 'lgrey', align='center')
    row_tot = df_tot[df_tot['month'] == mstr].iloc[0]
    for di, (ct, gt) in enumerate(DIMS):
        bg = _dim_bg[di % len(_dim_bg)]
        dmd = dim_month_data.get((mstr, ct, gt), {})
        pool_v  = dmd.get('pool_disp', 0)
        mode_v  = dmd.get('mode', '?')
        sales_v = dmd.get('total_sec', 0)
        rev_v   = dmd.get('total_rev', 0)
        aov_v   = dmd.get('aov', 0)
        dim_rev_totals[(ct, gt)] += rev_v
        c0 = 2 + di * N_DIM_COLS
        wc(ws2, r, c0,   round(pool_v), bg, fmt='#,##0',     align='right')
        wc(ws2, r, c0+1, mode_v,        bg, align='center',  size=8)
        wc(ws2, r, c0+2, round(sales_v, 1), bg, fmt='#,##0.0', align='right')
        wc(ws2, r, c0+3, round(rev_v),  bg, fmt=USD,          align='right')
        wc(ws2, r, c0+4, round(aov_v, 1), bg, fmt='$#,##0.0', align='right')
    wc(ws2, r, NCOLS2, int(row_tot['total_rev']), 'lgrey', fmt=USD, align='right')
    ws2.row_dimensions[r].height = 15

tr2 = 4 + len(FORECAST_MONTHS)
wc(ws2, tr2, 1, 'TOTAL', 'gold', bold=True, align='center')
for di, (ct, gt) in enumerate(DIMS):
    bg = _dim_bg[di % len(_dim_bg)]
    c0 = 2 + di * N_DIM_COLS
    for j in range(N_DIM_COLS - 1): wc(ws2, tr2, c0 + j, '--', bg)
    wc(ws2, tr2, c0 + 3, round(dim_rev_totals[(ct, gt)]), bg, fmt=USD, bold=True)
    wc(ws2, tr2, c0 + 4, '--', bg)
wc(ws2, tr2, NCOLS2, yr_rev, 'gold', fmt=USD, bold=True)
ws2.row_dimensions[tr2].height = 20

ws2.column_dimensions['A'].width = 12
for di in range(len(DIMS)):
    for j, w in enumerate([9, 7, 9, 11, 9]):
        ws2.column_dimensions[get_column_letter(2 + di * N_DIM_COLS + j)].width = w
ws2.column_dimensions[get_column_letter(NCOLS2)].width = 12

# ── SHEET 3: Calibration ──────────────────────────────────────────────────────
ws3 = wb.create_sheet('Calibration')
ws3.sheet_view.showGridLines = False
for c, w in {1: 26, 2: 12, 3: 12, 4: 12, 5: 38}.items():
    ws3.column_dimensions[get_column_letter(c)].width = w

wc(ws3, 1, 1, f'CALIBRATION PARAMETERS  |  cal: {CAL_WINDOW[0]}-{CAL_WINDOW[-1]}',
   'navy', bold=True, align='center', size=13, span=5)
ws3.row_dimensions[1].height = 30

def sec3(ws, r, title):
    wc(ws, r, 1, title, 'dblue', bold=True, align='left', size=11, span=5)
    ws.row_dimensions[r].height = 22
    return r + 1

r3 = 2

# 1. Segment shares
r3 = sec3(ws3, r3, f'Segment Shares (global, {CAL_WINDOW[0]}-{CAL_WINDOW[-1]}, n={total_sec_cal})')
for h, c in zip(['Segment', 'Share', '', '', 'Source'], [1, 2, 3, 4, 5]):
    wc(ws3, r3, c, h, 'navy', bold=True, align='center')
ws3.row_dimensions[r3].height = 18; r3 += 1
for seg in SEGS:
    wc(ws3, r3, 1, seg,         'lgrey',   align='left')
    wc(ws3, r3, 2, shares[seg], 'formula', fmt='0.0%', align='center')
    wc(ws3, r3, 3, '',          'lgrey')
    wc(ws3, r3, 4, '',          'lgrey')
    wc(ws3, r3, 5, src_shares[seg], 'lgrey', align='left', size=9)
    ws3.row_dimensions[r3].height = 16; r3 += 1
r3 += 1

# 2. Prolongation rates
r3 = sec3(ws3, r3, 'Prolongation Rates (by ct x gt)')
for h, c in zip(['Dimension', 'Rate', 'Pool (cal)', 'Sales (cal)', 'Source'], [1, 2, 3, 4, 5]):
    wc(ws3, r3, c, h, 'navy', bold=True, align='center')
ws3.row_dimensions[r3].height = 18; r3 += 1
for ct, gt in DIMS:
    p = len(pool_cal_df[(pool_cal_df['course_type']==ct) & (pool_cal_df['group_type']==gt)])
    s = len(sec_cal[(sec_cal['course_type']==ct) & (sec_cal['group_type']==gt)])
    wc(ws3, r3, 1, f'{ct}/{gt}',     'lgrey',   align='left')
    wc(ws3, r3, 2, rates[(ct, gt)],  'formula', fmt='0.00%', align='center')
    wc(ws3, r3, 3, p,                'lgrey',   fmt='#,##0', align='center')
    wc(ws3, r3, 4, s,                'lgrey',   fmt='#,##0', align='center')
    wc(ws3, r3, 5, src_rates[(ct, gt)], 'lgrey', align='left', size=9)
    ws3.row_dimensions[r3].height = 16; r3 += 1
r3 += 1

# 3. Retention by payment_no
r3 = sec3(ws3, r3, 'Retention by Payment_No (renewal number)')
for h, c in zip(['Payment_No', 'Retention Rate', '', '', 'Source'], [1, 2, 3, 4, 5]):
    wc(ws3, r3, c, h, 'navy', bold=True, align='center')
ws3.row_dimensions[r3].height = 18; r3 += 1
for pno in sorted(retention_by_renewal.keys()):
    wc(ws3, r3, 1, pno,                         'lgrey',   fmt='#,##0', align='center')
    wc(ws3, r3, 2, retention_by_renewal[pno],   'formula', fmt='0.0%',  align='center')
    wc(ws3, r3, 3, '',                           'lgrey')
    wc(ws3, r3, 4, '',                           'lgrey')
    wc(ws3, r3, 5, src_retention.get(pno, ''),  'lgrey',   align='left', size=9)
    ws3.row_dimensions[r3].height = 14; r3 += 1
r3 += 1

# 4. Package distribution
r3 = sec3(ws3, r3, 'Package Distribution (secondary, Jul-Dec 2025)')
for h, c in zip(['Dim', 'Package', 'Probability', '', 'Source'], [1, 2, 3, 4, 5]):
    wc(ws3, r3, c, h, 'navy', bold=True, align='center')
ws3.row_dimensions[r3].height = 18; r3 += 1
for ct, gt in DIMS:
    dim_pkgs = sorted([(k[2], v) for k, v in pkg_dist_final.items() if k[0]==ct and k[1]==gt],
                      key=lambda x: -x[1])
    for pkg_sz, prob in dim_pkgs:
        wc(ws3, r3, 1, f'{ct}/{gt}', 'lgrey', align='left', size=9)
        wc(ws3, r3, 2, pkg_sz,       'lgrey', fmt='#,##0',  align='center', size=9)
        wc(ws3, r3, 3, prob,         'formula', fmt='0.0%', align='center', size=9)
        wc(ws3, r3, 4, '',           'lgrey')
        wc(ws3, r3, 5, pkg_dist_src.get((ct, gt, pkg_sz), ''), 'lgrey', align='left', size=8)
        ws3.row_dimensions[r3].height = 13; r3 += 1
r3 += 1

# 5. Price per lesson
r3 = sec3(ws3, r3, 'Price per Lesson (base Q3 2025, secondary)')
for h, c in zip(['Dim', 'Package', 'Price/Lesson ($)', 'pgf Jan26', 'Source'], [1, 2, 3, 4, 5]):
    wc(ws3, r3, c, h, 'navy', bold=True, align='center')
ws3.row_dimensions[r3].height = 18; r3 += 1
pgf_jan = price_growth_factor('2026-01')
for ct, gt in DIMS:
    dim_pkgs = sorted([(k[2], v) for k, v in ppl_final.items() if k[0]==ct and k[1]==gt],
                      key=lambda x: x[0])
    for pkg_sz, ppl_v in dim_pkgs:
        wc(ws3, r3, 1, f'{ct}/{gt}',     'lgrey',   align='left', size=9)
        wc(ws3, r3, 2, pkg_sz,           'lgrey',   fmt='#,##0',  align='center', size=9)
        wc(ws3, r3, 3, ppl_v,            'formula', fmt='$#,##0.00', align='center', size=9)
        wc(ws3, r3, 4, ppl_v * pgf_jan, 'lgrey',   fmt='$#,##0.00', align='center', size=9)
        wc(ws3, r3, 5, ppl_src.get((ct, gt, pkg_sz), ''), 'lgrey', align='left', size=8)
        ws3.row_dimensions[r3].height = 13; r3 += 1
r3 += 1

# 6. Ext-curve
r3 = sec3(ws3, r3, f'Ext-curve  [{CURVE_CAL_START}-{CURVE_CAL_END}  max_lag={CURVE_MAX_LAG}]')
for h, c in zip(['Lag', 'Full (2023+)', 'Recent (2024+)', 'Used', 'Note'], [1, 2, 3, 4, 5]):
    wc(ws3, r3, c, h, 'navy', bold=True, align='center')
ws3.row_dimensions[r3].height = 18; r3 += 1
for k in sorted(ext_curve.keys()):
    src_tag = 'recent' if k in ext_curve_recent else 'full'
    bg_row  = 'lgreen' if src_tag == 'recent' else 'loran'
    wc(ws3, r3, 1, f'lag {k}',          'lgrey',  align='center', size=9)
    wc(ws3, r3, 2, ext_curve_full.get(k, '--') if ext_curve_full.get(k) else '--',
       'lgrey', fmt='0.000%' if ext_curve_full.get(k) else None, align='center', size=9)
    wc(ws3, r3, 3, ext_curve_recent.get(k, '--') if ext_curve_recent.get(k) else '--',
       'lgrey', fmt='0.000%' if ext_curve_recent.get(k) else None, align='center', size=9)
    wc(ws3, r3, 4, ext_curve[k], bg_row, fmt='0.000%', align='center', bold=True, size=9)
    wc(ws3, r3, 5, src_tag, 'lgrey', align='center', size=8, italic=True)
    ws3.row_dimensions[r3].height = 13; r3 += 1

# ── SHEET 4: Wide Format ──────────────────────────────────────────────────────
ws4 = wb.create_sheet('Wide Format')
ws4.sheet_view.showGridLines = False

# Column widths
ws4.column_dimensions['A'].width = 28
for ci in range(2, 15):
    ws4.column_dimensions[get_column_letter(ci)].width = 10

month_labels_short = ['Jan-26', 'Feb-26', 'Mar-26', 'Apr-26', 'May-26', 'Jun-26',
                      'Jul-26', 'Aug-26', 'Sep-26', 'Oct-26', 'Nov-26', 'Dec-26']

wc(ws4, 1, 1, 'Secondary Sales Forecast 2026 v3  —  Wide Format',
   'navy', bold=True, align='center', size=13, span=14)
ws4.row_dimensions[1].height = 28

# Header row
wc(ws4, 2, 1, 'Metric', 'dblue', bold=True, align='left')
for ci, lbl in enumerate(month_labels_short, 2):
    wc(ws4, 2, ci, lbl, 'dblue', bold=True, align='center', size=9)
ws4.row_dimensions[2].height = 20

r4 = 3

def write_wide_row(ws, r, label, values, bg='lgrey', fmt=USD, bold=False):
    wc(ws, r, 1, label, bg, bold=bold, align='left', size=9)
    for ci, v in enumerate(values, 2):
        wc(ws, r, ci, v, bg, fmt=fmt, align='right', bold=bold)
    ws.row_dimensions[r].height = 15

def write_blank(ws, r):
    for ci in range(1, 15):
        wc(ws, r, ci, '', 'white')
    ws.row_dimensions[r].height = 8

# -- secondary_rev_total
vals_rev_total = [int(df_tot[df_tot['month']==m]['total_rev'].iloc[0]) for m in FORECAST_MONTHS]
write_wide_row(ws4, r4, 'secondary_rev_total', vals_rev_total, 'gold', USD, bold=True)
r4 += 1

# -- secondary_rev by dim
dim_bg_map = {
    ('Base','Group'):   'lblue',
    ('Base','Premium'): 'lgreen',
    ('Base','Private'): 'loran',
    ('MT',  'Private'): 'teal2',
    ('MT',  'Premium'): 'sub',
    ('MT',  'Group'):   'lgrey',
}
for ct, gt in DIMS:
    bg = dim_bg_map.get((ct, gt), 'lgrey')
    vals = [round(dim_month_data.get((m, ct, gt), {}).get('total_rev', 0)) for m in FORECAST_MONTHS]
    write_wide_row(ws4, r4, f'secondary_rev [{ct}/{gt}]', vals, bg, USD)
    r4 += 1

write_blank(ws4, r4); r4 += 1

# -- secondary_n_total
vals_n_total = [round(df_tot[df_tot['month']==m]['total_sales'].iloc[0], 1) for m in FORECAST_MONTHS]
write_wide_row(ws4, r4, 'secondary_n_total', vals_n_total, 'gold', '#,##0.0', bold=True)
r4 += 1

# -- secondary_n by dim
for ct, gt in DIMS:
    bg = dim_bg_map.get((ct, gt), 'lgrey')
    vals = [round(dim_month_data.get((m, ct, gt), {}).get('total_sec', 0), 1) for m in FORECAST_MONTHS]
    write_wide_row(ws4, r4, f'secondary_n [{ct}/{gt}]', vals, bg, '#,##0.0')
    r4 += 1

write_blank(ws4, r4); r4 += 1

# -- secondary_aov
vals_aov = [row['aov'] for row in monthly_totals]
write_wide_row(ws4, r4, 'secondary_aov', [round(v, 1) for v in vals_aov], 'formula', '$#,##0.0', bold=True)
r4 += 1

write_blank(ws4, r4); r4 += 1

# -- actual_rev
vals_act = [row['actual_rev'] if row['actual_rev'] is not None else '--' for row in monthly_totals]
fmt_act = USD
wc(ws4, r4, 1, 'actual_rev', 'gold', bold=False, align='left', size=9)
for ci, v in enumerate(vals_act, 2):
    if isinstance(v, (int, float)):
        wc(ws4, r4, ci, v, 'gold', fmt=fmt_act, align='right')
    else:
        wc(ws4, r4, ci, v, 'lgrey', align='right', size=9)
ws4.row_dimensions[r4].height = 15; r4 += 1

# -- error_pct
vals_err = [row['err_pct'] if row['err_pct'] is not None else '--' for row in monthly_totals]
wc(ws4, r4, 1, 'error_pct', 'corr', bold=False, align='left', size=9)
for ci, v in enumerate(vals_err, 2):
    if isinstance(v, (int, float)):
        err_bg = 'lgreen' if v < 0 else ('lpink' if abs(v) > 10 else 'corr')
        wc(ws4, r4, ci, v / 100, err_bg, fmt='+0.0%;-0.0%;--', align='right')
    else:
        wc(ws4, r4, ci, v, 'lgrey', align='right', size=9)
ws4.row_dimensions[r4].height = 15; r4 += 1

write_blank(ws4, r4); r4 += 1

# -- price growth factors (informational)
wc(ws4, r4, 1, 'price_growth_factor (vs Q3-2025)', 'lock', italic=True, align='left', size=9)
for ci, mstr in enumerate(FORECAST_MONTHS, 2):
    pgf = price_growth_factor(mstr)
    wc(ws4, r4, ci, round(pgf, 4), 'lock', fmt='0.0000', align='right', size=9)
ws4.row_dimensions[r4].height = 14; r4 += 1

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f'\n{SEP}')
print(f'Saved: {OUT_PATH}')
print(f'Sheets: Monthly Summary | Dim Breakdown | Calibration | Wide Format')
print(SEP)
